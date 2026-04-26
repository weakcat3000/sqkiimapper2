from __future__ import annotations

from pathlib import Path
import subprocess

import pandas as pd
from openpyxl.utils import get_column_letter

from sp500_metrics_export import (
    BENCHMARK_TICKER,
    RISK_FREE_RATE,
    TRADING_DAYS_PER_YEAR,
    apply_header_style,
    auto_fit_columns,
    calculate_summary_metrics,
    download_market_data,
    extract_field,
    fetch_sp500_constituents,
    get_download_end_date,
)


DEFAULT_START_DATE = "2003-11-08"
DEFAULT_END_DATE = "2023-11-08"
OUTPUT_FILE = "reverse_engineering_analysis.xlsx"

STAGE1_MAX_20Y_ANNUALIZED_RETURN_RANK = 150
STAGE2_MIN_SHARPE_RATIO = 0.70
STAGE2_MIN_ANNUALIZED_RETURN = 0.15
STAGE2_MIN_YEARS_OF_HISTORY = 15
STAGE3_SHORTLIST_COUNT = 25
FINAL_SELECTION_COUNT = 18
STAGE4_ADDITIONAL_COUNT = 9
WEIGHT_SOURCE_ANNUALIZED = 3.0
WEIGHT_SHARPE = 4.0
WEIGHT_VOLATILITY = 1.0
KNOWN_TARGET_SYMBOLS = ["BKNG", "COST", "ISRG", "LLY", "MNST", "MSFT", "NVDA", "ODFL", "V"]

STRATEGIC_CATEGORY_RULES = [
    ("Travel", "Consumer Discretionary", "Hotels, Resorts & Cruise Lines"),
    ("Retail", "Consumer Staples", "Consumer Staples Merchandise Retail"),
    ("Medical Devices", "Health Care", "Health Care Equipment"),
    ("Pharmaceuticals", "Health Care", "Pharmaceuticals"),
    ("Beverages", "Consumer Staples", "Soft Drinks & Non-alcoholic Beverages"),
    ("Software", "Information Technology", "Systems Software"),
    ("Semiconductors", "Information Technology", "Semiconductors"),
    ("Transportation", "Industrials", "Cargo Ground Transportation"),
    ("Payment Systems", "Financials", "Transaction & Payment Processing Services"),
]

UNIVERSE_COLUMNS = [
    "symbol",
    "security",
    "sector",
    "sub_industry",
    "selection_category",
    "years_of_history",
    "return_1y",
    "annualized_return_since_start",
    "sharpe_ratio_since_start",
    "historical_volatility_since_start",
    "source_window_start_date_20y",
    "source_window_end_date_20y",
    "source_total_return_20y",
    "source_annualized_return_20y",
    "source_total_return_20y_rank",
    "source_annualized_return_20y_rank",
    "annualized_return_since_start_rank",
    "sharpe_ratio_since_start_rank",
    "historical_volatility_since_start_rank",
    "selection_score",
    "category_representative_rank",
    "final_selection_rank",
    "additional_finalist_rank",
    "in_stage1_long_term_winners",
    "in_stage2_quality_compounders",
    "in_stage3_diversified_candidates",
    "is_stage4_final_selection",
]

METRIC_DATA_COLUMNS = [
    "symbol",
    "history_start_date",
    "history_end_date",
    "years_of_history",
    "latest_price",
    "start_price_since_start",
    "total_return_since_start",
    "annualized_return_since_start",
    "price_1y_ago",
    "return_1y",
    "mean_excess_daily_return",
    "stdev_daily_return",
    "sharpe_ratio_since_start",
    "historical_volatility_since_start",
    "source_window_start_date_20y",
    "source_window_end_date_20y",
    "source_window_start_price_20y",
    "source_window_end_price_20y",
    "source_total_return_20y",
    "source_annualized_return_20y",
    "source_total_return_20y_rank",
    "source_annualized_return_20y_rank",
]

PERCENTAGE_COLUMNS = {
    "return_1y",
    "annualized_return_since_start",
    "historical_volatility_since_start",
    "source_total_return_20y",
    "source_annualized_return_20y",
    "total_return_since_start",
}

RATIO_COLUMNS = {"sharpe_ratio_since_start"}
YEARS_COLUMNS = {"years_of_history"}
CURRENCY_COLUMNS = {
    "latest_price",
    "start_price_since_start",
    "price_1y_ago",
    "source_window_start_price_20y",
    "source_window_end_price_20y",
}
DATE_COLUMNS = {
    "history_start_date",
    "history_end_date",
    "source_window_start_date_20y",
    "source_window_end_date_20y",
}


def map_selection_category(row: pd.Series) -> str:
    for label, sector, sub_industry in STRATEGIC_CATEGORY_RULES:
        if row["sector"] == sector and row["sub_industry"] == sub_industry:
            return label
    return ""


def build_parameters_sheet() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ["analysis_start_date", pd.Timestamp(DEFAULT_START_DATE), "", "input"],
            ["analysis_end_date", pd.Timestamp(DEFAULT_END_DATE), "", "input"],
            ["stage1_max_20y_annualized_return_rank", STAGE1_MAX_20Y_ANNUALIZED_RETURN_RANK, "", "stage_1"],
            ["stage2_min_sharpe_ratio_since_start", STAGE2_MIN_SHARPE_RATIO, "", "stage_2"],
            ["stage2_min_annualized_return_since_start", STAGE2_MIN_ANNUALIZED_RETURN, "", "stage_2"],
            ["stage2_min_years_of_history", STAGE2_MIN_YEARS_OF_HISTORY, "", "stage_2"],
            ["risk_free_rate_annual", RISK_FREE_RATE, "", "metric_input"],
            ["risk_free_rate_daily", f"=B8/{TRADING_DAYS_PER_YEAR}", "=B8/252", "metric_input"],
            ["one_year_lookback_date", "=EDATE(B3,-12)", "=EDATE(B3,-12)", "metric_input"],
            ["twenty_year_lookback_date", "=EDATE(B3,-240)", "=EDATE(B3,-240)", "metric_input"],
            ["stage3_shortlist_count", STAGE3_SHORTLIST_COUNT, "", "stage_3"],
            ["stage4_finalist_count", FINAL_SELECTION_COUNT, "", "stage_4"],
            ["stage4_additional_non_category_count", STAGE4_ADDITIONAL_COUNT, "", "stage_4"],
            ["weight_source_annualized_20y_rank", WEIGHT_SOURCE_ANNUALIZED, "", "scoring"],
            ["weight_sharpe_ratio_since_start_rank", WEIGHT_SHARPE, "", "scoring"],
            ["weight_historical_volatility_rank", WEIGHT_VOLATILITY, "", "scoring"],
        ],
        columns=["parameter", "value", "formula", "used_in_stage"],
    )


def build_universe_frame(constituents: pd.DataFrame) -> pd.DataFrame:
    universe = constituents[["symbol", "security", "sector", "sub_industry"]].copy()
    universe["selection_category"] = universe.apply(map_selection_category, axis=1)
    for column in UNIVERSE_COLUMNS:
        if column not in universe.columns:
            universe[column] = ""
    return universe[UNIVERSE_COLUMNS].sort_values("symbol").reset_index(drop=True)


def build_metric_data_frame(constituents: pd.DataFrame, available_tickers: set[str]) -> pd.DataFrame:
    rows = []
    for _, row in constituents.iterrows():
        base = {column: "" for column in METRIC_DATA_COLUMNS}
        base["symbol"] = row["symbol"]
        base["_yf_ticker"] = row["yf_ticker"] if row["yf_ticker"] in available_tickers else ""
        rows.append(base)
    return pd.DataFrame(rows)


def build_python_stage_frame(constituents: pd.DataFrame, universe: pd.DataFrame, summary_metrics: pd.DataFrame) -> pd.DataFrame:
    metrics = summary_metrics.rename(columns={"yf_ticker": "yf_ticker"}).copy()
    metrics["source_window_start_date_20y"] = metrics["history_start_date"]
    metrics["source_window_end_date_20y"] = metrics["history_end_date"]
    metrics["source_total_return_20y"] = metrics["total_return_since_start"]
    metrics["source_annualized_return_20y"] = metrics["annualized_return_since_start"]
    metrics["source_total_return_20y_rank"] = metrics["source_total_return_20y"].rank(
        ascending=False, method="min", na_option="bottom"
    )
    metrics["source_annualized_return_20y_rank"] = metrics["source_annualized_return_20y"].rank(
        ascending=False, method="min", na_option="bottom"
    )

    stage_frame = universe[
        [
            "symbol",
            "security",
            "sector",
            "sub_industry",
            "selection_category",
        ]
    ].merge(constituents[["symbol", "yf_ticker"]], on="symbol", how="left").merge(
        metrics[
            [
                "yf_ticker",
                "years_of_history",
                "return_1y",
                "annualized_return_since_start",
                "sharpe_ratio_since_start",
                "historical_volatility_since_start",
                "source_window_start_date_20y",
                "source_window_end_date_20y",
                "source_total_return_20y",
                "source_annualized_return_20y",
                "source_total_return_20y_rank",
                "source_annualized_return_20y_rank",
            ]
        ],
        on="yf_ticker",
        how="left",
    )

    stage_frame["annualized_return_since_start_rank"] = stage_frame["annualized_return_since_start"].rank(
        ascending=False,
        method="min",
        na_option="bottom",
    )
    stage_frame["sharpe_ratio_since_start_rank"] = stage_frame["sharpe_ratio_since_start"].rank(
        ascending=False,
        method="min",
        na_option="bottom",
    )
    stage_frame["historical_volatility_since_start_rank"] = stage_frame["historical_volatility_since_start"].rank(
        ascending=True,
        method="min",
        na_option="bottom",
    )
    stage_frame["selection_score"] = (
        WEIGHT_SOURCE_ANNUALIZED * stage_frame["source_annualized_return_20y_rank"]
        + WEIGHT_SHARPE * stage_frame["sharpe_ratio_since_start_rank"]
        + WEIGHT_VOLATILITY * stage_frame["historical_volatility_since_start_rank"]
    )

    stage_frame["in_stage1_long_term_winners"] = (
        stage_frame["source_annualized_return_20y_rank"] <= STAGE1_MAX_20Y_ANNUALIZED_RETURN_RANK
    ).fillna(False)

    stage_frame["in_stage2_quality_compounders"] = (
        stage_frame["in_stage1_long_term_winners"]
        & (stage_frame["sharpe_ratio_since_start"] >= STAGE2_MIN_SHARPE_RATIO)
        & (stage_frame["annualized_return_since_start"] >= STAGE2_MIN_ANNUALIZED_RETURN)
        & (stage_frame["years_of_history"] >= STAGE2_MIN_YEARS_OF_HISTORY)
    ).fillna(False)

    stage_frame["category_representative_rank"] = pd.NA
    category_ranked = (
        stage_frame.loc[stage_frame["in_stage2_quality_compounders"]]
        .sort_values(
            ["selection_category", "selection_score", "source_annualized_return_20y_rank", "symbol"],
            ascending=[True, True, True, True],
        )
        .copy()
    )
    category_ranked = category_ranked.loc[category_ranked["selection_category"].ne("")].copy()
    category_ranked["category_representative_rank"] = category_ranked.groupby("selection_category").cumcount() + 1
    stage_frame["final_selection_rank"] = pd.NA
    final_ranked = (
        stage_frame.loc[stage_frame["in_stage2_quality_compounders"]]
        .sort_values(
            ["selection_score", "source_annualized_return_20y_rank", "symbol"],
            ascending=[True, True, True],
        )
        .copy()
    )
    final_ranked["final_selection_rank"] = range(1, len(final_ranked) + 1)
    stage_frame.loc[category_ranked.index, "category_representative_rank"] = category_ranked["category_representative_rank"]
    stage_frame.loc[final_ranked.index, "final_selection_rank"] = final_ranked["final_selection_rank"]

    stage_frame["in_stage3_diversified_candidates"] = (
        stage_frame["in_stage2_quality_compounders"]
        & (
            stage_frame["final_selection_rank"].le(STAGE3_SHORTLIST_COUNT)
            | stage_frame["category_representative_rank"].eq(1)
        )
    ).fillna(False)

    stage_frame["additional_finalist_rank"] = pd.NA
    additional_ranked = (
        stage_frame.loc[
            stage_frame["in_stage3_diversified_candidates"] & ~stage_frame["category_representative_rank"].eq(1)
        ]
        .sort_values(
            ["selection_score", "source_annualized_return_20y_rank", "symbol"],
            ascending=[True, True, True],
        )
        .copy()
    )
    additional_ranked["additional_finalist_rank"] = range(1, len(additional_ranked) + 1)
    stage_frame.loc[additional_ranked.index, "additional_finalist_rank"] = additional_ranked["additional_finalist_rank"]

    stage_frame["is_stage4_final_selection"] = (
        stage_frame["category_representative_rank"].eq(1)
        | (
            stage_frame["in_stage3_diversified_candidates"]
            & stage_frame["additional_finalist_rank"].le(STAGE4_ADDITIONAL_COUNT)
        )
    ).fillna(False)
    return stage_frame


def build_stage_criteria(universe: pd.DataFrame) -> pd.DataFrame:
    last_row = len(universe) + 1
    cols = sheet_column_map(UNIVERSE_COLUMNS)
    return pd.DataFrame(
        [
            [
                0,
                "All S&P 500 Stocks",
                "Starting universe before any narrowing.",
                "Current S&P 500 constituent.",
                f'=COUNTA(universe!$A$2:$A${last_row})',
                "",
            ],
            [
                1,
                "Long-Term Winners",
                "Build the source pool from the full S&P 500.",
                f"20Y annualized-return rank <= {STAGE1_MAX_20Y_ANNUALIZED_RETURN_RANK}.",
                f'=COUNTIF(universe!${cols["in_stage1_long_term_winners"]}$2:${cols["in_stage1_long_term_winners"]}${last_row},TRUE)',
                "",
            ],
            [
                2,
                "Quality Compounders",
                "Keep only the strong long-term winners.",
                (
                    f"Stage 1 passed AND sharpe_ratio_since_start >= {STAGE2_MIN_SHARPE_RATIO:.2f} "
                    f"AND annualized_return_since_start >= {STAGE2_MIN_ANNUALIZED_RETURN:.0%} "
                    f"AND years_of_history >= {STAGE2_MIN_YEARS_OF_HISTORY}"
                ),
                f'=COUNTIF(universe!${cols["in_stage2_quality_compounders"]}$2:${cols["in_stage2_quality_compounders"]}${last_row},TRUE)',
                "",
            ],
            [
                3,
                "Scored Shortlist",
                "Take the best-scoring Stage 2 names while keeping the highest-ranked stock in each strategic category.",
                (
                    f"Stage 2 passed AND (final_selection_rank <= {STAGE3_SHORTLIST_COUNT} "
                    f"OR category_representative_rank = 1), "
                    f"where score = {WEIGHT_SOURCE_ANNUALIZED:g}*20Y annualized-return rank + "
                    f"{WEIGHT_SHARPE:g}*Sharpe rank + "
                    f"{WEIGHT_VOLATILITY:g}*volatility rank."
                ),
                f'=COUNTIF(universe!${cols["in_stage3_diversified_candidates"]}$2:${cols["in_stage3_diversified_candidates"]}${last_row},TRUE)',
                "",
            ],
            [
                4,
                "Finalists",
                "Keep all strategic-category representatives and add the next-best remaining names by score.",
                (
                    f"category_representative_rank = 1 "
                    f"OR (Stage 3 passed AND additional_finalist_rank <= {STAGE4_ADDITIONAL_COUNT}), "
                    f"for a final set of {FINAL_SELECTION_COUNT} stocks."
                ),
                f'=COUNTIF(universe!${cols["is_stage4_final_selection"]}$2:${cols["is_stage4_final_selection"]}${last_row},TRUE)',
                "",
            ],
        ],
        columns=[
            "stage_number",
            "stage_name",
            "why_this_stage_exists",
            "keep_if",
            "surviving_stocks",
            "notes",
        ],
    )


def sheet_column_map(columns: list[str]) -> dict[str, str]:
    return {column: get_column_letter(index + 1) for index, column in enumerate(columns)}


def build_stage_sheet_rows(selected_symbols: list[str], universe_row_map: dict[str, int], columns: list[str]) -> pd.DataFrame:
    rows = []
    for symbol in selected_symbols:
        row_number = universe_row_map[symbol]
        rows.append(
            {column: f"=universe!{get_column_letter(index)}{row_number}" for index, column in enumerate(columns, start=1)}
        )
    return pd.DataFrame(rows, columns=columns)


def build_validation_sheet() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "known_symbol": KNOWN_TARGET_SYMBOLS,
            "selected_by_model": [""] * len(KNOWN_TARGET_SYMBOLS),
            "final_selection_rank": [""] * len(KNOWN_TARGET_SYMBOLS),
            "sector": [""] * len(KNOWN_TARGET_SYMBOLS),
            "annualized_return_since_start": [""] * len(KNOWN_TARGET_SYMBOLS),
            "sharpe_ratio_since_start": [""] * len(KNOWN_TARGET_SYMBOLS),
        }
    )


def apply_number_formats(worksheet, columns: list[str]) -> None:
    header_map = sheet_column_map(columns)
    for name, letter in header_map.items():
        if name in PERCENTAGE_COLUMNS:
            number_format = "0.00%"
        elif name in RATIO_COLUMNS:
            number_format = "0.000"
        elif name in YEARS_COLUMNS:
            number_format = "0.00"
        elif name in CURRENCY_COLUMNS:
            number_format = '$#,##0.00'
        elif name in DATE_COLUMNS or name.endswith("_date"):
            number_format = "yyyy-mm-dd"
        else:
            continue

        for cell in worksheet[letter][1:]:
            cell.number_format = number_format


def format_header_label(name: str) -> str:
    label = name.replace("_", " ").title()
    replacements = {
        "20Y": "20Y",
        "1Y": "1Y",
        "Sp500": "SP500",
        "Sp 500": "SP 500",
        "Bkng": "BKNG",
        "Cost": "COST",
        "Isrg": "ISRG",
        "Lly": "LLY",
        "Mnst": "MNST",
        "Msft": "MSFT",
        "Nvda": "NVDA",
        "Odfl": "ODFL",
        "Ttm": "TTM",
        "Id": "ID",
    }
    for old, new in replacements.items():
        label = label.replace(old, new)
    return label


def prettify_headers(worksheet) -> None:
    for cell in worksheet[1]:
        if isinstance(cell.value, str):
            cell.value = format_header_label(cell.value)


def populate_metric_data_formulas(worksheet, price_data: pd.DataFrame, return_data: pd.DataFrame, metric_data: pd.DataFrame) -> None:
    price_last_row = len(price_data) + 1
    return_last_row = len(return_data) + 1
    metric_cols = sheet_column_map(METRIC_DATA_COLUMNS)

    for row_number, yf_ticker in enumerate(metric_data["_yf_ticker"], start=2):
        if not yf_ticker:
            continue

        price_col_idx = price_data.columns.get_loc(yf_ticker) + 1
        return_col_idx = return_data.columns.get_loc(yf_ticker) + 1
        price_letter = get_column_letter(price_col_idx)
        return_letter = get_column_letter(return_col_idx)

        price_range = f"price_data!${price_letter}$2:${price_letter}${price_last_row}"
        return_range = f"return_data!${return_letter}$2:${return_letter}${return_last_row}"
        date_range = f"price_data!$A$2:$A${price_last_row}"
        first_row_index = f'MATCH(TRUE,INDEX(ISNUMBER({price_range}),0),0)'
        last_row_index = f'LOOKUP(2,1/ISNUMBER({price_range}),ROW({price_range})-ROW(price_data!${price_letter}$2)+1)'

        worksheet[f"{metric_cols['history_start_date']}{row_number}"] = f'=IFERROR(INDEX({date_range},{first_row_index}),"")'
        worksheet[f"{metric_cols['history_end_date']}{row_number}"] = f'=IFERROR(INDEX({date_range},{last_row_index}),"")'
        worksheet[f"{metric_cols['years_of_history']}{row_number}"] = (
            f'=IFERROR(({metric_cols["history_end_date"]}{row_number}-{metric_cols["history_start_date"]}{row_number})/365.25,"")'
        )
        worksheet[f"{metric_cols['latest_price']}{row_number}"] = f'=IFERROR(LOOKUP(2,1/ISNUMBER({price_range}),{price_range}),"")'
        worksheet[f"{metric_cols['start_price_since_start']}{row_number}"] = f'=IFERROR(INDEX({price_range},{first_row_index}),"")'
        worksheet[f"{metric_cols['total_return_since_start']}{row_number}"] = (
            f'=IFERROR({metric_cols["latest_price"]}{row_number}/{metric_cols["start_price_since_start"]}{row_number}-1,"")'
        )
        worksheet[f"{metric_cols['annualized_return_since_start']}{row_number}"] = (
            f'=IFERROR(({metric_cols["latest_price"]}{row_number}/{metric_cols["start_price_since_start"]}{row_number})^(1/{metric_cols["years_of_history"]}{row_number})-1,"")'
        )
        worksheet[f"{metric_cols['price_1y_ago']}{row_number}"] = (
            f'=IFERROR(LOOKUP(2,1/(({date_range}<=parameters!$B$10)*ISNUMBER({price_range})),{price_range}),"")'
        )
        worksheet[f"{metric_cols['return_1y']}{row_number}"] = (
            f'=IFERROR({metric_cols["latest_price"]}{row_number}/{metric_cols["price_1y_ago"]}{row_number}-1,"")'
        )
        worksheet[f"{metric_cols['mean_excess_daily_return']}{row_number}"] = (
            f'=IFERROR(AVERAGE({return_range})-parameters!$B$9,"")'
        )
        worksheet[f"{metric_cols['stdev_daily_return']}{row_number}"] = f'=IFERROR(STDEV({return_range}),"")'
        worksheet[f"{metric_cols['sharpe_ratio_since_start']}{row_number}"] = (
            f'=IF(OR({metric_cols["mean_excess_daily_return"]}{row_number}="",{metric_cols["stdev_daily_return"]}{row_number}="",{metric_cols["stdev_daily_return"]}{row_number}=0),"",SQRT({TRADING_DAYS_PER_YEAR})*{metric_cols["mean_excess_daily_return"]}{row_number}/{metric_cols["stdev_daily_return"]}{row_number})'
        )
        worksheet[f"{metric_cols['historical_volatility_since_start']}{row_number}"] = (
            f'=IFERROR(SQRT({TRADING_DAYS_PER_YEAR})*{metric_cols["stdev_daily_return"]}{row_number},"")'
        )
        worksheet[f"{metric_cols['source_window_start_date_20y']}{row_number}"] = f"=B{row_number}"
        worksheet[f"{metric_cols['source_window_end_date_20y']}{row_number}"] = f"=C{row_number}"
        worksheet[f"{metric_cols['source_window_start_price_20y']}{row_number}"] = f"=F{row_number}"
        worksheet[f"{metric_cols['source_window_end_price_20y']}{row_number}"] = f"=E{row_number}"
        worksheet[f"{metric_cols['source_total_return_20y']}{row_number}"] = f"=G{row_number}"
        worksheet[f"{metric_cols['source_annualized_return_20y']}{row_number}"] = f"=H{row_number}"

    last_metric_row = len(metric_data) + 1
    for row_number in range(2, last_metric_row + 1):
        worksheet[f"{metric_cols['source_total_return_20y_rank']}{row_number}"] = (
            f'=IF(S{row_number}="","",RANK(S{row_number},$S$2:$S${last_metric_row},0))'
        )
        worksheet[f"{metric_cols['source_annualized_return_20y_rank']}{row_number}"] = (
            f'=IF(T{row_number}="","",RANK(T{row_number},$T$2:$T${last_metric_row},0))'
        )


def populate_universe_formulas(worksheet, universe: pd.DataFrame) -> None:
    universe_cols = sheet_column_map(UNIVERSE_COLUMNS)
    metric_cols = sheet_column_map(METRIC_DATA_COLUMNS)
    last_row = len(universe) + 1

    linked_columns = {
        "years_of_history": "years_of_history",
        "return_1y": "return_1y",
        "annualized_return_since_start": "annualized_return_since_start",
        "sharpe_ratio_since_start": "sharpe_ratio_since_start",
        "historical_volatility_since_start": "historical_volatility_since_start",
        "source_window_start_date_20y": "source_window_start_date_20y",
        "source_window_end_date_20y": "source_window_end_date_20y",
        "source_total_return_20y": "source_total_return_20y",
        "source_annualized_return_20y": "source_annualized_return_20y",
        "source_total_return_20y_rank": "source_total_return_20y_rank",
        "source_annualized_return_20y_rank": "source_annualized_return_20y_rank",
    }

    for row_number in range(2, last_row + 1):
        for universe_column, metric_column in linked_columns.items():
            worksheet[f"{universe_cols[universe_column]}{row_number}"] = (
                f'=IFERROR(INDEX(metric_data!${metric_cols[metric_column]}:${metric_cols[metric_column]},MATCH($A{row_number},metric_data!$A:$A,0)),"")'
            )
        annual_rank = universe_cols["annualized_return_since_start_rank"]
        sharpe_rank = universe_cols["sharpe_ratio_since_start_rank"]
        vol_rank = universe_cols["historical_volatility_since_start_rank"]
        score_col = universe_cols["selection_score"]
        category_col = universe_cols["selection_category"]
        category_rank_col = universe_cols["category_representative_rank"]
        final_rank_col = universe_cols["final_selection_rank"]
        additional_rank_col = universe_cols["additional_finalist_rank"]
        stage1_col = universe_cols["in_stage1_long_term_winners"]
        stage2_col = universe_cols["in_stage2_quality_compounders"]
        stage3_col = universe_cols["in_stage3_diversified_candidates"]
        stage4_col = universe_cols["is_stage4_final_selection"]
        ann_return_col = universe_cols["annualized_return_since_start"]
        sharpe_col = universe_cols["sharpe_ratio_since_start"]
        years_col = universe_cols["years_of_history"]
        ann20_rank_col = universe_cols["source_annualized_return_20y_rank"]

        worksheet[f"{annual_rank}{row_number}"] = (
            f'=IF(${ann_return_col}{row_number}="","",RANK(${ann_return_col}{row_number},${ann_return_col}$2:${ann_return_col}${last_row},0))'
        )
        worksheet[f"{sharpe_rank}{row_number}"] = (
            f'=IF(${sharpe_col}{row_number}="","",RANK(${sharpe_col}{row_number},${sharpe_col}$2:${sharpe_col}${last_row},0))'
        )
        worksheet[f"{vol_rank}{row_number}"] = (
            f'=IF(${universe_cols["historical_volatility_since_start"]}{row_number}="","",RANK(${universe_cols["historical_volatility_since_start"]}{row_number},${universe_cols["historical_volatility_since_start"]}$2:${universe_cols["historical_volatility_since_start"]}${last_row},1))'
        )
        worksheet[f"{score_col}{row_number}"] = (
            f'=IF(COUNT(${ann20_rank_col}{row_number},${sharpe_rank}{row_number},${vol_rank}{row_number})<3,"",parameters!$B$15*${ann20_rank_col}{row_number}+parameters!$B$16*${sharpe_rank}{row_number}+parameters!$B$17*${vol_rank}{row_number})'
        )
        worksheet[f"{stage1_col}{row_number}"] = (
            f'=AND(COUNT(${ann20_rank_col}{row_number})>0,${ann20_rank_col}{row_number}<=parameters!$B$4)'
        )
        worksheet[f"{stage2_col}{row_number}"] = (
            f'=AND(${stage1_col}{row_number}=TRUE,${sharpe_col}{row_number}>=parameters!$B$5,${ann_return_col}{row_number}>=parameters!$B$6,${years_col}{row_number}>=parameters!$B$7)'
        )
        worksheet[f"{category_rank_col}{row_number}"] = (
            f'=IF(OR(${stage2_col}{row_number}=FALSE,${category_col}{row_number}=""),"",1+COUNTIFS(${category_col}$2:${category_col}${last_row},${category_col}{row_number},${stage2_col}$2:${stage2_col}${last_row},TRUE,${score_col}$2:${score_col}${last_row},"<"&${score_col}{row_number})+COUNTIFS(${category_col}$2:${category_col}${last_row},${category_col}{row_number},${stage2_col}$2:${stage2_col}${last_row},TRUE,${score_col}$2:${score_col}${last_row},${score_col}{row_number},${ann20_rank_col}$2:${ann20_rank_col}${last_row},"<"&${ann20_rank_col}{row_number})+COUNTIFS(${category_col}$2:${category_col}${last_row},${category_col}{row_number},${stage2_col}$2:${stage2_col}${last_row},TRUE,${score_col}$2:${score_col}${last_row},${score_col}{row_number},${ann20_rank_col}$2:${ann20_rank_col}${last_row},${ann20_rank_col}{row_number},$A$2:$A${last_row},"<"&$A{row_number}))'
        )
        worksheet[f"{final_rank_col}{row_number}"] = (
            f'=IF(${stage2_col}{row_number}=FALSE,"",1+COUNTIFS(${stage2_col}$2:${stage2_col}${last_row},TRUE,${score_col}$2:${score_col}${last_row},"<"&${score_col}{row_number})+COUNTIFS(${stage2_col}$2:${stage2_col}${last_row},TRUE,${score_col}$2:${score_col}${last_row},${score_col}{row_number},${ann20_rank_col}$2:${ann20_rank_col}${last_row},"<"&${ann20_rank_col}{row_number})+COUNTIFS(${stage2_col}$2:${stage2_col}${last_row},TRUE,${score_col}$2:${score_col}${last_row},${score_col}{row_number},${ann20_rank_col}$2:${ann20_rank_col}${last_row},${ann20_rank_col}{row_number},$A$2:$A${last_row},"<"&$A{row_number}))'
        )
        worksheet[f"{stage3_col}{row_number}"] = (
            f'=AND(${stage2_col}{row_number}=TRUE,OR(${final_rank_col}{row_number}<=parameters!$B$12,${category_rank_col}{row_number}=1))'
        )
        worksheet[f"{additional_rank_col}{row_number}"] = (
            f'=IF(OR(${stage3_col}{row_number}=FALSE,${category_rank_col}{row_number}=1),"",1+COUNTIFS(${stage3_col}$2:${stage3_col}${last_row},TRUE,${category_rank_col}$2:${category_rank_col}${last_row},"<>1",${score_col}$2:${score_col}${last_row},"<"&${score_col}{row_number})+COUNTIFS(${stage3_col}$2:${stage3_col}${last_row},TRUE,${category_rank_col}$2:${category_rank_col}${last_row},"<>1",${score_col}$2:${score_col}${last_row},${score_col}{row_number},${ann20_rank_col}$2:${ann20_rank_col}${last_row},"<"&${ann20_rank_col}{row_number})+COUNTIFS(${stage3_col}$2:${stage3_col}${last_row},TRUE,${category_rank_col}$2:${category_rank_col}${last_row},"<>1",${score_col}$2:${score_col}${last_row},${score_col}{row_number},${ann20_rank_col}$2:${ann20_rank_col}${last_row},${ann20_rank_col}{row_number},$A$2:$A${last_row},"<"&$A{row_number}))'
        )
        worksheet[f"{stage4_col}{row_number}"] = (
            f'=OR(${category_rank_col}{row_number}=1,AND(${stage3_col}{row_number}=TRUE,${additional_rank_col}{row_number}<=parameters!$B$14))'
        )


def populate_validation_formulas(worksheet, validation_sheet: pd.DataFrame) -> None:
    universe_cols = sheet_column_map(UNIVERSE_COLUMNS)
    last_row = len(validation_sheet) + 1
    for row_number in range(2, last_row + 1):
        worksheet[f"B{row_number}"] = f'=COUNTIF(stage4_finalists!$A:$A,$A{row_number})>0'
        worksheet[f"C{row_number}"] = f'=IFERROR(INDEX(universe!${universe_cols["final_selection_rank"]}:${universe_cols["final_selection_rank"]},MATCH($A{row_number},universe!$A:$A,0)),"")'
        worksheet[f"D{row_number}"] = f'=IFERROR(INDEX(universe!$C:$C,MATCH($A{row_number},universe!$A:$A,0)),"")'
        worksheet[f"E{row_number}"] = f'=IFERROR(INDEX(universe!${universe_cols["annualized_return_since_start"]}:${universe_cols["annualized_return_since_start"]},MATCH($A{row_number},universe!$A:$A,0)),"")'
        worksheet[f"F{row_number}"] = f'=IFERROR(INDEX(universe!${universe_cols["sharpe_ratio_since_start"]}:${universe_cols["sharpe_ratio_since_start"]},MATCH($A{row_number},universe!$A:$A,0)),"")'


def force_excel_recalc_and_save(workbook_path: Path) -> None:
    escaped_path = str(workbook_path.resolve()).replace("'", "''")
    command = rf"""
$path = '{escaped_path}'
$excel = $null
$workbook = $null
try {{
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $workbook = $excel.Workbooks.Open($path, $false, $false)
    $excel.Calculation = -4105
    $excel.CalculateFullRebuild()
    $workbook.Save()
}}
catch {{
}}
finally {{
    if ($workbook -ne $null) {{ $workbook.Close($true) }}
    if ($excel -ne $null) {{ $excel.Quit() }}
}}
"""
    subprocess.run(
        ["powershell", "-NoProfile", "-Command", command],
        check=False,
        capture_output=True,
        text=True,
    )


def export_to_excel(
    parameters: pd.DataFrame,
    stage_criteria: pd.DataFrame,
    universe: pd.DataFrame,
    metric_data: pd.DataFrame,
    price_data: pd.DataFrame,
    return_data: pd.DataFrame,
    stage1_sheet: pd.DataFrame,
    stage2_sheet: pd.DataFrame,
    stage3_sheet: pd.DataFrame,
    stage4_sheet: pd.DataFrame,
    validation_sheet: pd.DataFrame,
    missing_tickers: pd.DataFrame,
    output_path: Path,
) -> Path:
    candidate_paths = [output_path, output_path.with_stem(f"{output_path.stem}_updated")]

    for candidate_path in candidate_paths:
        try:
            with pd.ExcelWriter(candidate_path, engine="openpyxl") as writer:
                parameters.to_excel(writer, sheet_name="parameters", index=False)
                stage_criteria.to_excel(writer, sheet_name="stage_criteria", index=False)
                universe.to_excel(writer, sheet_name="universe", index=False)
                stage1_sheet.to_excel(writer, sheet_name="stage1_long_term_winners", index=False)
                stage2_sheet.to_excel(writer, sheet_name="stage2_quality_compounders", index=False)
                stage3_sheet.to_excel(writer, sheet_name="stage3_scored_shortlist", index=False)
                stage4_sheet.to_excel(writer, sheet_name="stage4_finalists", index=False)
                validation_sheet.to_excel(writer, sheet_name="validation", index=False)
                metric_data.drop(columns="_yf_ticker").to_excel(writer, sheet_name="metric_data", index=False)
                price_data.to_excel(writer, sheet_name="price_data", index=False)
                return_data.to_excel(writer, sheet_name="return_data", index=False)
                missing_tickers.to_excel(writer, sheet_name="missing_tickers", index=False)

                workbook = writer.book
                workbook.calculation.calcMode = "auto"
                workbook.calculation.fullCalcOnLoad = True
                workbook.calculation.forceFullCalc = True

                for worksheet in writer.sheets.values():
                    prettify_headers(worksheet)
                    apply_header_style(worksheet)
                    worksheet.freeze_panes = "A2"

                populate_metric_data_formulas(writer.sheets["metric_data"], price_data, return_data, metric_data)
                populate_universe_formulas(writer.sheets["universe"], universe)
                populate_validation_formulas(writer.sheets["validation"], validation_sheet)

                for sheet_name, columns in {
                    "universe": universe.columns.tolist(),
                    "metric_data": metric_data.drop(columns="_yf_ticker").columns.tolist(),
                    "stage1_long_term_winners": stage1_sheet.columns.tolist(),
                    "stage2_quality_compounders": stage2_sheet.columns.tolist(),
                    "stage3_scored_shortlist": stage3_sheet.columns.tolist(),
                    "stage4_finalists": stage4_sheet.columns.tolist(),
                    "validation": validation_sheet.columns.tolist(),
                }.items():
                    apply_number_formats(writer.sheets[sheet_name], columns)

                auto_fit_columns(writer.sheets["parameters"])
                auto_fit_columns(writer.sheets["stage_criteria"])
                auto_fit_columns(writer.sheets["universe"])
                auto_fit_columns(writer.sheets["stage1_long_term_winners"])
                auto_fit_columns(writer.sheets["stage2_quality_compounders"])
                auto_fit_columns(writer.sheets["stage3_scored_shortlist"])
                auto_fit_columns(writer.sheets["stage4_finalists"])
                auto_fit_columns(writer.sheets["validation"])
                auto_fit_columns(writer.sheets["metric_data"])
                auto_fit_columns(writer.sheets["missing_tickers"])

                for data_sheet in ("price_data", "return_data"):
                    auto_fit_columns(writer.sheets[data_sheet])
                    for cell in writer.sheets[data_sheet]["A"][1:]:
                        cell.number_format = "yyyy-mm-dd"

                for cell in writer.sheets["parameters"]["B"][1:]:
                    if cell.row in {2, 3, 10, 11}:
                        cell.number_format = "yyyy-mm-dd"
                    elif cell.row in {5, 8, 9, 15, 16, 17, 18}:
                        cell.number_format = "0.000"
                    elif cell.row == 6:
                        cell.number_format = "0.00%"
                    elif cell.row in {4, 7, 12, 13, 14}:
                        cell.number_format = "0"

            force_excel_recalc_and_save(candidate_path)
            return candidate_path
        except PermissionError:
            continue

    raise PermissionError(
        f"Could not write to {output_path.name} or {output_path.with_stem(f'{output_path.stem}_updated').name}. "
        "Please close the workbook and run the script again."
    )


def main() -> None:
    output_path = Path(__file__).with_name(OUTPUT_FILE)

    constituents = fetch_sp500_constituents()
    tickers = constituents["yf_ticker"].tolist()
    raw_market_data = download_market_data(
        tickers + [BENCHMARK_TICKER],
        start=DEFAULT_START_DATE,
        end=get_download_end_date(DEFAULT_END_DATE),
    )

    adjusted_close = extract_field(raw_market_data, "Adj Close").ffill()
    close_prices = extract_field(raw_market_data, "Close").ffill()
    dividends = extract_field(raw_market_data, "Dividends").fillna(0.0)

    if BENCHMARK_TICKER not in adjusted_close.columns:
        raise RuntimeError(f"Benchmark ticker {BENCHMARK_TICKER} was not returned by yfinance.")

    available_tickers = set(adjusted_close.columns) - {BENCHMARK_TICKER}
    summary_metrics = calculate_summary_metrics(
        prices=adjusted_close,
        close_prices=close_prices,
        dividends=dividends,
        benchmark_ticker=BENCHMARK_TICKER,
        analysis_start_date=DEFAULT_START_DATE,
        analysis_end_date=DEFAULT_END_DATE,
    )

    universe = build_universe_frame(constituents)
    metric_data = build_metric_data_frame(constituents, available_tickers)
    python_stage_frame = build_python_stage_frame(constituents, universe, summary_metrics)

    price_data = adjusted_close.drop(columns=BENCHMARK_TICKER, errors="ignore").loc[:, sorted(available_tickers)].reset_index()
    price_data = price_data.rename(columns={price_data.columns[0]: "date"})
    return_data = adjusted_close.drop(columns=BENCHMARK_TICKER, errors="ignore").pct_change(fill_method=None)
    return_data = return_data.loc[:, sorted(available_tickers)].reset_index()
    return_data = return_data.rename(columns={return_data.columns[0]: "date"})

    universe_row_map = {symbol: index + 2 for index, symbol in enumerate(universe["symbol"])}
    stage1_sheet = build_stage_sheet_rows(
        python_stage_frame.loc[python_stage_frame["in_stage1_long_term_winners"], "symbol"].sort_values().tolist(),
        universe_row_map,
        UNIVERSE_COLUMNS,
    )
    stage2_sheet = build_stage_sheet_rows(
        python_stage_frame.loc[python_stage_frame["in_stage2_quality_compounders"], "symbol"].sort_values().tolist(),
        universe_row_map,
        UNIVERSE_COLUMNS,
    )
    stage3_sheet = build_stage_sheet_rows(
        python_stage_frame.loc[python_stage_frame["in_stage3_diversified_candidates"], "symbol"].sort_values().tolist(),
        universe_row_map,
        UNIVERSE_COLUMNS,
    )
    stage4_sheet = build_stage_sheet_rows(
        python_stage_frame.loc[python_stage_frame["is_stage4_final_selection"], "symbol"].sort_values().tolist(),
        universe_row_map,
        UNIVERSE_COLUMNS,
    )
    validation_sheet = build_validation_sheet()

    exported_path = export_to_excel(
        parameters=build_parameters_sheet(),
        stage_criteria=build_stage_criteria(universe),
        universe=universe,
        metric_data=metric_data,
        price_data=price_data,
        return_data=return_data,
        stage1_sheet=stage1_sheet,
        stage2_sheet=stage2_sheet,
        stage3_sheet=stage3_sheet,
        stage4_sheet=stage4_sheet,
        validation_sheet=validation_sheet,
        missing_tickers=constituents.loc[
            ~constituents["yf_ticker"].isin(available_tickers), ["symbol", "yf_ticker", "security"]
        ].reset_index(drop=True),
        output_path=output_path,
    )

    print(f"Exported workbook to: {exported_path}")


if __name__ == "__main__":
    main()
