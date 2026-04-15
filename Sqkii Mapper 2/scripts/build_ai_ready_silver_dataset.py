from __future__ import annotations

import csv
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[2]
EXPORTS_DIR = ROOT / "exports"
SOURCE_CSV = EXPORTS_DIR / "sqkii_silver_combined_reviewed.csv"
OUT_CSV = EXPORTS_DIR / "sqkii_silver_ai_ready.csv"
OUT_JSONL = EXPORTS_DIR / "sqkii_silver_ai_ready.jsonl"
OUT_XLSX = EXPORTS_DIR / "sqkii_silver_ai_ready.xlsx"
PUBLIC_JSON = ROOT / "Sqkii Mapper 2" / "public" / "silver-ai-dataset.json"


POSITION_PATTERNS = [
    ("in_between", ["in between", "between"]),
    ("inside", ["inside"]),
    ("behind", ["behind"]),
    ("under", ["under"]),
    ("below", ["below"]),
    ("beside", ["beside"]),
    ("on_top", ["on top of", "on top"]),
    ("above", ["above"]),
    ("at_base", ["at the base of", "at the planter edge", "at the walkway edge", "at the base edge"]),
    ("in_corner", ["in the corner", "corner gap"]),
    ("on", ["on "]),
]


OBJECT_KEYWORDS = [
    ("bench", ["bench", "seat", "slat", "armrest"]),
    ("pipe", ["pipe", "drainpipe"]),
    ("wall", ["wall", "retaining wall"]),
    ("column", ["column", "pillar"]),
    ("railing", ["railing"]),
    ("fence", ["fence", "gate"]),
    ("panel", ["panel", "plinth"]),
    ("signboard", ["signboard"]),
    ("lamp_post", ["lamp post", "lamp post", "lamppost", "pole base", "pole"]),
    ("utility_cover", ["manhole", "utility cover", "utility base", "utility box", "utility cap", "cover frame", "cover"]),
    ("drain", ["drain", "grate", "drainage slot"]),
    ("kerb", ["kerb", "curb"]),
    ("planter", ["planter", "planting bed"]),
    ("hedge", ["hedge", "bush", "shrub", "groundcover"]),
    ("tree", ["tree", "trunk", "roots", "root"]),
    ("grass", ["grass"]),
    ("soil", ["soil", "mulch"]),
    ("leaves", ["leaf", "leaves", "leaf litter"]),
    ("rock", ["rock", "rocks", "stone", "stones", "pebbles", "gravel", "rubble"]),
    ("brick", ["brick", "bricks"]),
    ("tile", ["tile", "tiles", "tiled"]),
    ("paver", ["paver", "pavers", "stepping stone", "pavement", "walkway", "path"]),
    ("concrete", ["concrete", "slab", "footing", "ledge", "step", "stairs", "stair"]),
    ("wood", ["wood", "wooden", "plank", "planks"]),
    ("playground", ["playground", "slide", "tunnel"]),
]


SURFACE_KEYWORDS = [
    ("concrete", ["concrete", "slab", "footing", "ledge", "kerb", "curb"]),
    ("soil", ["soil", "mulch", "bare soil"]),
    ("grass", ["grass"]),
    ("leaf_litter", ["leaf", "leaves"]),
    ("metal", ["metal", "railing", "panel", "pipe", "grate"]),
    ("brick", ["brick"]),
    ("wood", ["wood", "wooden", "plank"]),
    ("rock", ["rock", "stone", "gravel", "pebbles", "rubble"]),
    ("tile", ["tile", "tiled", "paver", "pavement", "walkway", "path"]),
]


def normalize(text: str) -> str:
    return re.sub(r"\s+", " ", text.strip().lower())


def find_tags(text: str, patterns: list[tuple[str, list[str]]], limit: int = 2) -> list[str]:
    matches: list[tuple[int, str]] = []
    for tag, variants in patterns:
        positions = [text.find(variant) for variant in variants if variant in text]
        positions = [pos for pos in positions if pos >= 0]
        if positions:
            matches.append((min(positions), tag))
    matches.sort(key=lambda item: item[0])
    ordered: list[str] = []
    for _, tag in matches:
        if tag not in ordered:
            ordered.append(tag)
        if len(ordered) >= limit:
            break
    return ordered


def derive_environment(row: dict[str, str], objects: list[str]) -> str:
    text = normalize(f"{row['spot_type']} {row['spot_description']}")
    if "playground" in objects:
        return "playground"
    if "bench" in objects:
        return "bench_zone"
    if any(obj in objects for obj in ["utility_cover", "drain", "pipe"]):
        return "utility_edge"
    if any(obj in objects for obj in ["tree", "hedge", "planter"]):
        return "landscape_edge"
    if any(obj in objects for obj in ["wall", "column", "panel", "fence", "railing", "signboard", "lamp_post"]):
        return "structure_edge"
    if any(word in text for word in ["walkway", "pavement", "path", "paver", "tile"]):
        return "walkway_edge"
    return "mixed_edge"


def derive_concealment(row: dict[str, str], positions: list[str], objects: list[str]) -> str:
    text = normalize(f"{row['spot_type']} {row['spot_description']}")
    if any(word in text for word in ["crack", "seam", "gap", "groove", "crevice", "slot"]):
        return "seam_gap"
    if any(word in text for word in ["hole", "opening"]):
        return "hole_opening"
    if "under" in positions or "below" in positions:
        return "underside"
    if "behind" in positions:
        return "behind_object"
    if "inside" in positions:
        return "inside_object"
    if any(obj in objects for obj in ["leaves", "hedge", "grass"]):
        return "natural_cover"
    return "edge_hide"


def derive_surface(text: str) -> str:
    tags = find_tags(text, SURFACE_KEYWORDS, limit=3)
    if not tags:
        return "mixed"
    if len(tags) == 1:
        return tags[0]
    return "+".join(tags[:2])


def derive_search_priority(environment: str, concealment: str, confidence: str) -> int:
    score = 3
    if concealment in {"seam_gap", "hole_opening", "underside", "inside_object"}:
        score = 5
    elif concealment in {"behind_object", "edge_hide"}:
        score = 4
    elif concealment == "natural_cover":
        score = 3

    if environment in {"utility_edge", "bench_zone", "structure_edge"}:
        score = min(5, score + 0)
    elif environment in {"landscape_edge", "walkway_edge"}:
        score = max(3, score - 0)

    if confidence == "low":
        score = max(2, score - 2)
    elif confidence == "medium":
        score = max(3, score - 1)
    return score


def build_search_text(row: dict[str, str], positions: list[str], objects: list[str], environment: str, concealment: str) -> str:
    parts = [
        f"{row['campaign']} silver coin",
        f"coin {row['coin_number']}",
        f"spot type {row['spot_type'].replace('_', ' ')}",
        f"positions {' '.join(positions) if positions else 'edge'}",
        f"objects {' '.join(objects) if objects else 'mixed'}",
        f"environment {environment.replace('_', ' ')}",
        f"concealment {concealment.replace('_', ' ')}",
        row["spot_description"],
    ]
    return ". ".join(parts)


def build_instruction(positions: list[str], objects: list[str], concealment: str) -> str:
    position = humanize_position(positions[0]) if positions else "at the edge of"
    obj = objects[0].replace("_", " ") if objects else "structure"
    if concealment == "seam_gap":
        return f"Check the seam or narrow gap {position} the {obj} before scanning open ground."
    if concealment == "underside":
        return f"Check underneath the {obj} first, then inspect the nearest edge and floor seam."
    if concealment == "behind_object":
        return f"Check behind the {obj} and along the base where debris collects."
    if concealment == "inside_object":
        return f"Check inside the {obj} opening or recess before sweeping nearby grass or soil."
    if concealment == "natural_cover":
        return f"Check under leaves, roots, or groundcover beside the {obj} before searching open pavement."
    return f"Start {position} the {obj} and inspect the nearest edge before checking open ground."


def humanize_position(position: str) -> str:
    mapping = {
        "in_between": "in between",
        "inside": "inside",
        "behind": "behind",
        "under": "under",
        "below": "below",
        "beside": "beside",
        "on_top": "on top of",
        "above": "above",
        "at_base": "at the base of",
        "in_corner": "in the corner of",
        "on": "on",
    }
    return mapping.get(position, position.replace("_", " "))


def confidence_score(confidence: str) -> float:
    return {"high": 1.0, "medium": 0.7, "low": 0.4}.get(confidence, 0.5)


def read_rows() -> list[dict[str, str]]:
    with SOURCE_CSV.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def enrich_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    enriched: list[dict[str, str]] = []
    for row in rows:
        source_text = normalize(f"{row['spot_type'].replace('_', ' ')}. {row['spot_description']}")
        positions = find_tags(source_text, POSITION_PATTERNS, limit=2)
        objects = find_tags(source_text, OBJECT_KEYWORDS, limit=3)
        environment = derive_environment(row, objects)
        concealment = derive_concealment(row, positions, objects)
        surface = derive_surface(source_text)
        priority = derive_search_priority(environment, concealment, row["confidence"])
        enriched.append(
            {
                "record_id": f"{row['campaign'].lower()}_{int(row['coin_number']):03d}",
                "campaign": row["campaign"],
                "coin_number": row["coin_number"],
                "short_id": row["short_id"],
                "spot_type": row["spot_type"],
                "spot_description": row["spot_description"],
                "confidence": row["confidence"],
                "confidence_score": f"{confidence_score(row['confidence']):.1f}",
                "review_basis": row["review_basis"],
                "primary_position": positions[0] if positions else "",
                "secondary_position": positions[1] if len(positions) > 1 else "",
                "primary_object": objects[0] if objects else "",
                "secondary_object": objects[1] if len(objects) > 1 else "",
                "tertiary_object": objects[2] if len(objects) > 2 else "",
                "environment_type": environment,
                "surface_type": surface,
                "concealment_type": concealment,
                "search_priority": str(priority),
                "search_instruction": build_instruction(positions, objects, concealment),
                "search_text": build_search_text(row, positions, objects, environment, concealment),
            }
        )
    return enriched


def write_csv(rows: list[dict[str, str]]) -> None:
    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_jsonl(rows: list[dict[str, str]]) -> None:
    with OUT_JSONL.open("w", encoding="utf-8") as f:
        for row in rows:
            payload = {
                "id": row["record_id"],
                "input_text": row["search_text"],
                "metadata": {
                    "campaign": row["campaign"],
                    "coin_number": int(row["coin_number"]),
                    "short_id": row["short_id"],
                    "spot_type": row["spot_type"],
                    "confidence": row["confidence"],
                    "confidence_score": float(row["confidence_score"]),
                    "primary_position": row["primary_position"],
                    "primary_object": row["primary_object"],
                    "environment_type": row["environment_type"],
                    "surface_type": row["surface_type"],
                    "concealment_type": row["concealment_type"],
                    "search_priority": int(row["search_priority"]),
                },
            }
            f.write(json.dumps(payload, ensure_ascii=True) + "\n")


def write_workbook(rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "AI Ready"
    ws.append(list(rows[0].keys()))
    for row in rows:
        ws.append([row[key] for key in rows[0].keys()])

    summary = wb.create_sheet("Summary")
    summary.append(["metric", "value"])
    summary.append(["reviewed_rows", len(rows)])
    campaign_counts = Counter(row["campaign"] for row in rows)
    for campaign, count in sorted(campaign_counts.items()):
        summary.append([f"campaign_{campaign}", count])

    priority_counts = Counter(row["search_priority"] for row in rows)
    for priority, count in sorted(priority_counts.items(), key=lambda item: int(item[0])):
        summary.append([f"search_priority_{priority}", count])

    for campaign in sorted(campaign_counts):
        sheet = wb.create_sheet(campaign)
        campaign_rows = [row for row in rows if row["campaign"] == campaign]
        sheet.append(list(rows[0].keys()))
        for row in campaign_rows:
            sheet.append([row[key] for key in rows[0].keys()])

    wb.save(OUT_XLSX)


def write_public_json(rows: list[dict[str, str]]) -> None:
    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "rowCount": len(rows),
        "campaigns": dict(sorted(Counter(row["campaign"] for row in rows).items())),
        "rows": rows,
    }
    with PUBLIC_JSON.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=True, indent=2)


def main() -> None:
    rows = read_rows()
    enriched = enrich_rows(rows)
    write_csv(enriched)
    write_jsonl(enriched)
    write_workbook(enriched)
    write_public_json(enriched)
    print(f"Wrote {OUT_CSV}")
    print(f"Wrote {OUT_JSONL}")
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {PUBLIC_JSON}")
    print(f"Rows: {len(enriched)}")


if __name__ == "__main__":
    main()
