from __future__ import annotations

import argparse
import json
import math
import os
import random
import re
import statistics
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
import requests


EMBED_LEN = 12
MIN_PREFIX_STEPS = 10
K_NEIGHBORS = 8
API_TABLE = "coin_history_archive"
FORMULA_VERSION = 2


@dataclass
class Step:
    lat: float
    lng: float
    radius_m: float
    timestamp_ms: Optional[int]


@dataclass
class RecordMetrics:
    start_radius_m: float
    floor_radius_m: float
    total_duration_s: float
    total_drift_m: float
    mean_step_distance_m: float
    shrink_rate_norm: float


@dataclass
class Record:
    entry_id: str
    room_code: str
    coin_label: str
    created_at: str
    exact_lat: float
    exact_lng: float
    steps: List[Step]
    metrics: RecordMetrics
    family_id: str = ""


@dataclass
class Sample:
    sample_id: str
    entry_id: str
    room_code: str
    coin_label: str
    prefix_steps: List[Step]
    prefix_len: int
    total_steps: int
    output_offset_m: Tuple[float, float]
    output_norm: Tuple[float, float]
    progress_to_floor: float
    remaining_progress: float
    total_duration_ratio: float
    remaining_duration_s: float
    profile: Dict[str, float]
    embedding: List[float]
    signature: List[float]
    exact_coords: Tuple[float, float]
    record_metrics: RecordMetrics
    family_id: str
    public_offset_m: Tuple[float, float]
    residual_offset_m: Tuple[float, float]
    next_step_offset_m: Tuple[float, float]
    next_radius_delta_norm: float
    next_turn_flag: float
    remaining_steps: int


@dataclass
class FeatureStats:
    means: List[float]
    stds: List[float]


def parse_args() -> argparse.Namespace:
    project_root = Path(__file__).resolve().parents[1]
    public_dir = project_root / "public"
    parser = argparse.ArgumentParser(
        description="Fit a browser-usable shrink predictor formula from archived exact-ended coins."
    )
    parser.add_argument("--room", default="", help="Optional room filter, e.g. silver")
    parser.add_argument("--limit", type=int, default=1000, help="Max archived records to fetch")
    parser.add_argument("--min-prefix-steps", type=int, default=MIN_PREFIX_STEPS)
    parser.add_argument("--target-error-m", type=float, default=200.0)
    parser.add_argument("--target-confidence-m", type=float, default=500.0)
    parser.add_argument("--time-budget-sec", type=float, default=180.0)
    parser.add_argument("--max-iterations", type=int, default=4000)
    parser.add_argument("--max-stale-iterations", type=int, default=600)
    parser.add_argument(
        "--max-prefix-samples-per-coin",
        type=int,
        default=0,
        help="Optional cap on prefix samples per coin using normalized progress anchors plus turns/resets. 0 keeps all prefixes.",
    )
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument(
        "--excel-path",
        default="",
        help="Optional path to a shrink-log Excel workbook. If provided, training uses the workbook instead of Supabase.",
    )
    parser.add_argument(
        "--output-formula",
        default=str(public_dir / "shrink_formula.latest.json"),
        help="Where to write the website-usable formula JSON",
    )
    parser.add_argument(
        "--output-report",
        default=str(public_dir / "shrink_formula.report.json"),
        help="Where to write the fitter report JSON",
    )
    parser.add_argument(
        "--env-file",
        default=str(project_root / ".env"),
        help="Path to the website .env file with Supabase credentials",
    )
    return parser.parse_args()


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip("\"").strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def require_env(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def mean(values: Sequence[float]) -> float:
    finite = [float(v) for v in values if math.isfinite(float(v))]
    if not finite:
        return float("nan")
    return sum(finite) / len(finite)


def weighted_mean(values: Sequence[float], weights: Sequence[float]) -> float:
    pairs = [
        (float(v), max(0.0, float(w)))
        for v, w in zip(values, weights)
        if math.isfinite(float(v)) and math.isfinite(float(w))
    ]
    total_weight = sum(weight for _, weight in pairs)
    if total_weight <= 0:
        return mean(values)
    return sum(value * weight for value, weight in pairs) / total_weight


def weighted_percentile(values: Sequence[float], weights: Sequence[float], percentile: float) -> float:
    pairs = sorted(
        [
            (float(v), max(0.0, float(w)))
            for v, w in zip(values, weights)
            if math.isfinite(float(v)) and math.isfinite(float(w))
        ],
        key=lambda item: item[0],
    )
    if not pairs:
        return float("nan")
    total_weight = sum(weight for _, weight in pairs)
    if total_weight <= 0:
        return pairs[-1][0]
    threshold = clamp(percentile, 0.0, 1.0) * total_weight
    acc = 0.0
    for value, weight in pairs:
        acc += weight
        if acc >= threshold:
            return value
    return pairs[-1][0]


def timestamp_ms(raw: Any) -> Optional[int]:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        value = int(raw)
        if value > 10_000_000_000:
            return value
        if value > 0:
            return value * 1000
        return None
    text = str(raw).strip()
    if not text:
        return None
    try:
        if text.endswith("Z"):
            dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
        else:
            dt = datetime.fromisoformat(text)
        return int(dt.timestamp() * 1000)
    except Exception:
        return None


def meters_between(a: Tuple[float, float], b: Tuple[float, float]) -> float:
    east, north = signed_offset_m(a, b)
    return math.hypot(east, north)


def signed_offset_m(origin: Tuple[float, float], target: Tuple[float, float]) -> Tuple[float, float]:
    meters_per_deg_lat = 111_320.0
    meters_per_deg_lng = 111_320.0 * math.cos(math.radians(origin[0]))
    east = (target[1] - origin[1]) * meters_per_deg_lng
    north = (target[0] - origin[0]) * meters_per_deg_lat
    return east, north


def destination(origin: Tuple[float, float], east_m: float, north_m: float) -> Tuple[float, float]:
    meters_per_deg_lat = 111_320.0
    meters_per_deg_lng = 111_320.0 * math.cos(math.radians(origin[0]))
    lat = origin[0] + (north_m / meters_per_deg_lat)
    lng = origin[1] + (east_m / max(1e-6, meters_per_deg_lng))
    return lat, lng


def wrap_angle_rad(value: float) -> float:
    while value <= -math.pi:
        value += 2 * math.pi
    while value > math.pi:
        value -= 2 * math.pi
    return value


def resample_series(values: Sequence[float], out_len: int = EMBED_LEN) -> List[float]:
    if out_len <= 0:
        return []
    if not values:
        return [0.0] * out_len
    if len(values) == 1:
        return [float(values[0])] * out_len
    result: List[float] = []
    last_index = len(values) - 1
    for i in range(out_len):
        pos = (i / max(1, out_len - 1)) * last_index
        left = int(math.floor(pos))
        right = min(last_index, left + 1)
        ratio = pos - left
        left_v = float(values[left])
        right_v = float(values[right])
        result.append((left_v * (1 - ratio)) + (right_v * ratio))
    return result


def normalize_weight_map(weights: Dict[str, float], fallback: Dict[str, float]) -> Dict[str, float]:
    merged = {**fallback, **(weights or {})}
    total = sum(max(0.0, float(value)) for value in merged.values())
    if total <= 0:
        total = sum(max(0.0, float(value)) for value in fallback.values()) or 1.0
        merged = dict(fallback)
    return {key: max(0.0, float(value)) / total for key, value in merged.items()}


def shift_weights(base: Dict[str, float], deltas: Dict[str, float], fallback: Dict[str, float]) -> Dict[str, float]:
    merged = {**fallback, **(base or {})}
    for key, delta in (deltas or {}).items():
        merged[key] = float(merged.get(key, 0.0)) + float(delta)
    return normalize_weight_map(merged, fallback)


def coerce_step(raw: Dict[str, Any]) -> Optional[Step]:
    if not isinstance(raw, dict):
        return None
    lat = raw.get("lat", raw.get("latitude"))
    lng = raw.get("lng", raw.get("longitude", raw.get("lon")))
    radius = raw.get("radiusMeters", raw.get("radius_m", raw.get("radius")))
    if lat is None or lng is None or radius is None:
        return None
    try:
        step = Step(
            lat=float(lat),
            lng=float(lng),
            radius_m=max(1.0, float(radius)),
            timestamp_ms=timestamp_ms(raw.get("ts", raw.get("timestamp", raw.get("created_at")))),
        )
    except Exception:
        return None
    if not math.isfinite(step.lat) or not math.isfinite(step.lng) or not math.isfinite(step.radius_m):
        return None
    return step


def normalize_steps(raw_steps: Sequence[Dict[str, Any]]) -> List[Step]:
    steps = [step for step in (coerce_step(item) for item in (raw_steps or [])) if step]
    steps.sort(key=lambda item: (item.timestamp_ms if item.timestamp_ms is not None else 2**62))
    return steps


def build_sequence_profile(steps: Sequence[Step]) -> Dict[str, float]:
    if not steps:
        return {}
    start = steps[0]
    end = steps[-1]
    start_radius = max(1.0, start.radius_m)
    radii = [step.radius_m for step in steps]
    drift_series = [meters_between((start.lat, start.lng), (step.lat, step.lng)) / start_radius for step in steps]
    step_distances = [0.0]
    for prev, current in zip(steps, steps[1:]):
        step_distances.append(meters_between((prev.lat, prev.lng), (current.lat, current.lng)) / start_radius)
    start_ts = start.timestamp_ms
    end_ts = end.timestamp_ms
    duration_s = 0.0
    if start_ts is not None and end_ts is not None and end_ts >= start_ts:
        duration_s = (end_ts - start_ts) / 1000.0
    shrink_amount = max(0.0, start_radius - end.radius_m)
    bundle_profile = build_direction_bundle_profile(steps)
    return {
        "stepCount": float(len(steps)),
        "startRadiusKm": start_radius / 1000.0,
        "currentRadiusKm": end.radius_m / 1000.0,
        "radiusRatio": end.radius_m / start_radius,
        "shrinkFraction": shrink_amount / start_radius,
        "driftRatio": drift_series[-1] if drift_series else 0.0,
        "meanStepDistanceRatio": mean(step_distances) if step_distances else 0.0,
        "durationNorm": duration_s / max(60.0, start_radius),
        "shrinkRateNorm": shrink_amount / max(1.0, duration_s, start_radius),
        "radiusStdRatio": statistics.pstdev(r / start_radius for r in radii) if len(radii) > 1 else 0.0,
        **bundle_profile,
    }


FAMILY_PROFILE_KEYS = [
    "startRadiusKm",
    "shrinkFraction",
    "meanStepDistanceRatio",
    "durationNorm",
    "shrinkRateNorm",
    "radiusStdRatio",
    "bundleCountRatio",
    "meanBundleRunRatio",
    "maxBundleRunRatio",
    "finalBundleRunRatio",
    "turnDensity",
    "headingChangeNorm",
    "finalHeadingEastNorm",
    "finalHeadingNorthNorm",
]


def build_family_vector(profile: Dict[str, float]) -> List[float]:
    return [float(profile.get(key, 0.0)) for key in FAMILY_PROFILE_KEYS]


def standardize_vectors(vectors: Sequence[Sequence[float]]) -> Tuple[List[List[float]], FeatureStats]:
    stats = build_feature_stats(vectors)
    return [standardize_vector(vector, stats) for vector in vectors], stats


def choose_family_count(record_count: int) -> int:
    if record_count <= 10:
        return 2
    if record_count <= 24:
        return 3
    if record_count <= 48:
        return 4
    return 5


def nearest_centroid(vector: Sequence[float], centroids: Sequence[Sequence[float]]) -> int:
    return min(range(len(centroids)), key=lambda index: distance(vector, centroids[index]))


def build_family_model(records: Sequence[Record], seed: int = 42) -> Dict[str, Any]:
    if not records:
        return {
            "keys": FAMILY_PROFILE_KEYS,
            "means": [],
            "stds": [],
            "centroids": [],
            "summaries": [],
        }

    raw_vectors = [build_family_vector(build_sequence_profile(record.steps)) for record in records]
    standardized, stats = standardize_vectors(raw_vectors)
    k = min(choose_family_count(len(records)), len(records))
    rng = random.Random(seed)
    init_indices = list(range(len(records)))
    rng.shuffle(init_indices)
    centroids = [list(standardized[index]) for index in init_indices[:k]]
    assignments = [0] * len(records)

    for _ in range(24):
        changed = False
        for index, vector in enumerate(standardized):
            family_idx = nearest_centroid(vector, centroids)
            if family_idx != assignments[index]:
                assignments[index] = family_idx
                changed = True
        for family_idx in range(k):
            members = [standardized[index] for index, assigned in enumerate(assignments) if assigned == family_idx]
            if members:
                centroids[family_idx] = [
                    mean([member[col] for member in members])
                    for col in range(len(centroids[family_idx]))
                ]
            else:
                replacement = standardized[rng.randrange(len(standardized))]
                centroids[family_idx] = list(replacement)
                changed = True
        if not changed:
            break

    summaries = []
    for family_idx in range(k):
        members = [record for record, assigned in zip(records, assignments) if assigned == family_idx]
        family_id = f"family-{family_idx + 1}"
        for record in members:
            record.family_id = family_id
        start_radii = [record.metrics.start_radius_m for record in members]
        floor_radii = [record.metrics.floor_radius_m for record in members]
        durations = [record.metrics.total_duration_s for record in members]
        summaries.append({
            "familyId": family_id,
            "count": len(members),
            "avgStartRadiusM": mean(start_radii) if start_radii else 0.0,
            "avgSmallestPublicCircleSizeM": mean(floor_radii) if floor_radii else 0.0,
            "avgDurationHours": (mean(durations) / 3600.0) if durations else 0.0,
            "centroid": centroids[family_idx],
            "labels": [record.coin_label for record in members[:8]],
        })

    return {
        "keys": FAMILY_PROFILE_KEYS,
        "means": stats.means,
        "stds": stats.stds,
        "centroids": centroids,
        "summaries": summaries,
    }


def classify_family(profile: Dict[str, float], family_model: Dict[str, Any]) -> Dict[str, Any]:
    centroids = family_model.get("centroids") or []
    if not centroids:
        return {"familyId": "", "scores": {}, "vector": []}
    raw_vector = build_family_vector(profile)
    vector = standardize_vector(raw_vector, FeatureStats(
        means=list(family_model.get("means") or []),
        stds=list(family_model.get("stds") or []),
    ))
    distances = [distance(vector, centroid) for centroid in centroids]
    inv = [1.0 / max(1e-6, value + 0.25) for value in distances]
    total = sum(inv) or 1.0
    scores = {
        (family_model.get("summaries") or [{}])[index].get("familyId", f"family-{index + 1}"): inv[index] / total
        for index in range(len(inv))
    }
    family_id = max(scores.items(), key=lambda item: item[1])[0] if scores else ""
    return {"familyId": family_id, "scores": scores, "vector": vector}


def build_normalized_signature(steps: Sequence[Step]) -> List[Dict[str, float]]:
    if not steps:
        return []
    start = steps[0]
    start_radius = max(1.0, start.radius_m)
    signature = []
    prev = None
    for step in steps:
        east, north = signed_offset_m((start.lat, start.lng), (step.lat, step.lng))
        step_distance = meters_between((prev.lat, prev.lng), (step.lat, step.lng)) if prev else 0.0
        signature.append({
            "radiusRatio": step.radius_m / start_radius,
            "eastRatio": east / start_radius,
            "northRatio": north / start_radius,
            "driftRatio": math.hypot(east, north) / start_radius,
            "stepDistanceRatio": step_distance / start_radius,
        })
        prev = step
    return signature


def build_direction_bundle_profile(steps: Sequence[Step]) -> Dict[str, float]:
    move_count = max(0, len(steps) - 1)
    if move_count <= 0:
        return {
            "bundleCountRatio": 0.0,
            "meanBundleRunRatio": 0.0,
            "maxBundleRunRatio": 0.0,
            "finalBundleRunRatio": 0.0,
            "turnDensity": 0.0,
            "headingChangeNorm": 0.0,
            "finalHeadingEastNorm": 0.0,
            "finalHeadingNorthNorm": 0.0,
        }

    bundle_lengths: List[int] = []
    heading_changes: List[float] = []
    current_bundle_len = 0
    prev_heading: Optional[float] = None
    final_heading_east = 0.0
    final_heading_north = 0.0
    threshold_rad = math.radians(26.0)

    for prev, current in zip(steps, steps[1:]):
        east, north = signed_offset_m((prev.lat, prev.lng), (current.lat, current.lng))
        dist = math.hypot(east, north)
        if dist <= 1e-6:
            heading = prev_heading if prev_heading is not None else 0.0
        else:
            heading = math.atan2(north, east)
        angle_delta = 0.0 if prev_heading is None else abs(wrap_angle_rad(heading - prev_heading))
        if prev_heading is None or angle_delta <= threshold_rad:
            current_bundle_len += 1
        else:
            bundle_lengths.append(max(1, current_bundle_len))
            current_bundle_len = 1
            heading_changes.append(angle_delta / math.pi)
        prev_heading = heading
        final_heading_east = math.cos(heading)
        final_heading_north = math.sin(heading)

    bundle_lengths.append(max(1, current_bundle_len))
    bundle_count = len(bundle_lengths)
    mean_bundle = mean(bundle_lengths) if bundle_lengths else 0.0
    max_bundle = max(bundle_lengths) if bundle_lengths else 0.0
    final_bundle = bundle_lengths[-1] if bundle_lengths else 0.0
    return {
        "bundleCountRatio": bundle_count / max(1.0, move_count),
        "meanBundleRunRatio": mean_bundle / max(1.0, move_count),
        "maxBundleRunRatio": max_bundle / max(1.0, move_count),
        "finalBundleRunRatio": final_bundle / max(1.0, move_count),
        "turnDensity": max(0.0, bundle_count - 1) / max(1.0, move_count),
        "headingChangeNorm": mean(heading_changes) if heading_changes else 0.0,
        "finalHeadingEastNorm": final_heading_east,
        "finalHeadingNorthNorm": final_heading_north,
    }


def build_bundle_signature_series(steps: Sequence[Step]) -> List[Dict[str, float]]:
    move_count = max(0, len(steps) - 1)
    if move_count <= 0:
        return []

    series: List[Dict[str, float]] = []
    current_bundle_len = 0
    prev_heading: Optional[float] = None
    threshold_rad = math.radians(26.0)

    for prev, current in zip(steps, steps[1:]):
        east, north = signed_offset_m((prev.lat, prev.lng), (current.lat, current.lng))
        dist = math.hypot(east, north)
        if dist <= 1e-6:
            heading = prev_heading if prev_heading is not None else 0.0
        else:
            heading = math.atan2(north, east)
        angle_delta = 0.0 if prev_heading is None else abs(wrap_angle_rad(heading - prev_heading))
        turn_flag = 0.0
        if prev_heading is None or angle_delta <= threshold_rad:
            current_bundle_len += 1
        else:
            current_bundle_len = 1
            turn_flag = 1.0
        prev_heading = heading
        series.append({
            "headingCos": math.cos(heading),
            "headingSin": math.sin(heading),
            "bundleRunRatio": current_bundle_len / max(1.0, move_count),
            "turnFlag": turn_flag,
        })
    return series


def build_signature_vector(steps: Sequence[Step], out_len: int = EMBED_LEN) -> List[float]:
    signature = build_normalized_signature(steps)
    bundle_series = build_bundle_signature_series(steps)
    if not signature:
        return [0.0] * (out_len * 9)
    return [
        *resample_series([item["radiusRatio"] for item in signature], out_len),
        *resample_series([item["eastRatio"] for item in signature], out_len),
        *resample_series([item["northRatio"] for item in signature], out_len),
        *resample_series([item["driftRatio"] for item in signature], out_len),
        *resample_series([item["stepDistanceRatio"] for item in signature], out_len),
        *resample_series([item["headingCos"] for item in bundle_series], out_len),
        *resample_series([item["headingSin"] for item in bundle_series], out_len),
        *resample_series([item["bundleRunRatio"] for item in bundle_series], out_len),
        *resample_series([item["turnFlag"] for item in bundle_series], out_len),
    ]


def build_embedding_from_steps(steps: Sequence[Step], out_len: int = EMBED_LEN) -> List[float]:
    signature = build_normalized_signature(steps)
    if not signature:
        return [0.0] * (out_len * 9)
    return build_signature_vector(steps, out_len)


def build_feature_stats(vectors: Sequence[Sequence[float]]) -> FeatureStats:
    if not vectors:
        return FeatureStats(means=[], stds=[])
    width = max(len(vector) for vector in vectors)
    means: List[float] = []
    stds: List[float] = []
    for index in range(width):
        column = [float(vector[index]) for vector in vectors if index < len(vector)]
        avg = mean(column)
        variance = mean([(value - avg) ** 2 for value in column])
        means.append(avg if math.isfinite(avg) else 0.0)
        stds.append(max(1e-6, math.sqrt(variance) if math.isfinite(variance) else 1.0))
    return FeatureStats(means=means, stds=stds)


def standardize_vector(vector: Sequence[float], stats: FeatureStats) -> List[float]:
    return [
        (float(vector[index]) - stats.means[index]) / max(1e-6, stats.stds[index])
        for index in range(min(len(vector), len(stats.means)))
    ]


def distance(a: Sequence[float], b: Sequence[float], indices: Optional[Sequence[int]] = None) -> float:
    if indices is None:
        size = min(len(a), len(b))
        return math.sqrt(sum((float(a[i]) - float(b[i])) ** 2 for i in range(size)))
    return math.sqrt(sum((float(a[i]) - float(b[i])) ** 2 for i in indices if i < len(a) and i < len(b)))


def build_record_metrics(steps: Sequence[Step], exact_coords: Tuple[float, float]) -> RecordMetrics:
    start = steps[0]
    end = steps[-1]
    total_duration_s = 0.0
    if start.timestamp_ms is not None and end.timestamp_ms is not None and end.timestamp_ms >= start.timestamp_ms:
        total_duration_s = (end.timestamp_ms - start.timestamp_ms) / 1000.0
    drift = meters_between((end.lat, end.lng), exact_coords)
    step_distances = [
        meters_between((prev.lat, prev.lng), (current.lat, current.lng))
        for prev, current in zip(steps, steps[1:])
    ]
    shrink_amount = max(0.0, start.radius_m - end.radius_m)
    shrink_rate_norm = shrink_amount / max(1.0, total_duration_s, start.radius_m)
    return RecordMetrics(
        start_radius_m=max(1.0, start.radius_m),
        floor_radius_m=max(1.0, end.radius_m),
        total_duration_s=total_duration_s,
        total_drift_m=drift,
        mean_step_distance_m=mean(step_distances) if step_distances else 0.0,
        shrink_rate_norm=shrink_rate_norm,
    )


def fetch_exact_records(supabase_url: str, supabase_key: str, limit: int, room: str = "") -> List[Dict[str, Any]]:
    url = f"{supabase_url.rstrip('/')}/rest/v1/{API_TABLE}"
    params = {
        "select": "id,room_code,coin_label,lifecycle,exact_lat,exact_lng,exact_note,created_at",
        "exact_lat": "not.is.null",
        "exact_lng": "not.is.null",
        "order": "created_at.desc",
        "limit": str(limit),
    }
    if room:
        params["room_code"] = f"eq.{room}"
    response = requests.get(
        url,
        params=params,
        headers={
            "apikey": supabase_key,
            "Authorization": f"Bearer {supabase_key}",
        },
        timeout=60,
    )
    response.raise_for_status()
    return response.json()


def parse_true_location(raw: Any) -> Optional[Tuple[float, float]]:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    parts = re.findall(r"[-+]?\d*\.?\d+", text)
    if len(parts) < 2:
        return None
    try:
        return float(parts[0]), float(parts[1])
    except Exception:
        return None


def load_excel_records(workbook_path: Path, min_prefix_steps: int) -> List[Record]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    if "Shrinks" not in wb.sheetnames:
        raise RuntimeError(f"Workbook does not contain a 'Shrinks' sheet: {workbook_path}")
    ws = wb["Shrinks"]
    block_width = 8
    grouped_rows: Dict[str, List[Dict[str, Any]]] = {}

    for row in ws.iter_rows(values_only=True):
        values = list(row)
        for start in range(0, len(values), block_width):
            block = values[start : start + block_width]
            if len(block) < 7:
                continue
            coin_id, coin_label, shrink_time, lat, lng, radius, true_location = block[:7]
            if coin_id == "Coin ID" or coin_id is None:
                continue
            try:
                lat_f = float(lat)
                lng_f = float(lng)
                radius_f = float(radius)
            except Exception:
                continue
            grouped_rows.setdefault(str(coin_id).strip(), []).append(
                {
                    "coin_label": str(coin_label or "").strip(),
                    "shrink_time": shrink_time,
                    "lat": lat_f,
                    "lng": lng_f,
                    "radius_m": radius_f,
                    "true_location": parse_true_location(true_location),
                }
            )

    records: List[Record] = []
    for coin_id, rows in grouped_rows.items():
        rows.sort(key=lambda item: timestamp_ms(item["shrink_time"]) or 2**62)
        exact_candidates = [item["true_location"] for item in rows if item["true_location"]]
        if not exact_candidates:
            continue
        steps = [
            Step(
                lat=float(item["lat"]),
                lng=float(item["lng"]),
                radius_m=max(1.0, float(item["radius_m"])),
                timestamp_ms=timestamp_ms(item["shrink_time"]),
            )
            for item in rows
        ]
        if len(steps) < min_prefix_steps:
            continue
        exact_lat, exact_lng = exact_candidates[-1]
        coin_label = rows[0]["coin_label"] or coin_id
        records.append(
            Record(
                entry_id=coin_id,
                room_code="excel",
                coin_label=coin_label,
                created_at=str(rows[0]["shrink_time"] or "").strip(),
                exact_lat=exact_lat,
                exact_lng=exact_lng,
                steps=steps,
                metrics=build_record_metrics(steps, (exact_lat, exact_lng)),
            )
        )
    return records


def build_records(raw_entries: Sequence[Dict[str, Any]], min_prefix_steps: int) -> List[Record]:
    records: List[Record] = []
    for entry in raw_entries:
        try:
            exact_lat = float(entry["exact_lat"])
            exact_lng = float(entry["exact_lng"])
        except Exception:
            continue
        steps = normalize_steps(entry.get("lifecycle") or [])
        if len(steps) < min_prefix_steps:
            continue
        metrics = build_record_metrics(steps, (exact_lat, exact_lng))
        records.append(
            Record(
                entry_id=str(entry.get("id", "")),
                room_code=str(entry.get("room_code", "")).strip().lower(),
                coin_label=str(entry.get("coin_label", "")).strip() or f"Coin {entry.get('id', '')}",
                created_at=str(entry.get("created_at", "")).strip(),
                exact_lat=exact_lat,
                exact_lng=exact_lng,
                steps=steps,
                metrics=metrics,
            )
        )
    return records


def select_prefix_end_indices(record: Record, min_prefix_steps: int, max_prefix_samples_per_coin: int) -> List[int]:
    available = list(range(min_prefix_steps, len(record.steps) + 1))
    if max_prefix_samples_per_coin <= 0 or len(available) <= max_prefix_samples_per_coin:
        return available

    start_radius = max(1.0, record.metrics.start_radius_m)
    shrinkable_radius = max(1.0, record.metrics.start_radius_m - record.metrics.floor_radius_m)
    threshold_rad = math.radians(26.0)
    progress_by_end: Dict[int, float] = {}
    event_indices: set[int] = set()
    prev_heading: Optional[float] = None

    for end in available:
        step = record.steps[end - 1]
        progress_by_end[end] = clamp(
            (start_radius - step.radius_m) / shrinkable_radius,
            0.0,
            1.0,
        )

    for idx in range(1, len(record.steps)):
        end = idx + 1
        if end < min_prefix_steps:
            continue
        prev_step = record.steps[idx - 1]
        step = record.steps[idx]
        if step.radius_m > prev_step.radius_m + 0.5:
            event_indices.add(end)
        east, north = signed_offset_m((prev_step.lat, prev_step.lng), (step.lat, step.lng))
        if math.hypot(east, north) > 1e-6:
            heading = math.atan2(north, east)
            if prev_heading is not None:
                delta = abs(wrap_angle_rad(heading - prev_heading))
                if delta > threshold_rad:
                    event_indices.add(end)
            prev_heading = heading

    selected = {available[0], available[-1], *event_indices}
    if len(selected) > max_prefix_samples_per_coin:
        sorted_selected = sorted(selected, key=lambda end: progress_by_end.get(end, 0.0))
        targets = [
            i / max(1, max_prefix_samples_per_coin - 1)
            for i in range(max_prefix_samples_per_coin)
        ]
        reduced: set[int] = set()
        for target in targets:
            choice = min(sorted_selected, key=lambda end: abs(progress_by_end.get(end, 0.0) - target))
            reduced.add(choice)
        selected = reduced

    remaining_slots = max_prefix_samples_per_coin - len(selected)
    if remaining_slots > 0:
        targets = [
            i / max(1, remaining_slots - 1)
            for i in range(remaining_slots)
        ] if remaining_slots > 1 else [0.5]
        candidates = [end for end in available if end not in selected]
        for target in targets:
            if not candidates:
                break
            choice = min(candidates, key=lambda end: abs(progress_by_end.get(end, 0.0) - target))
            selected.add(choice)
            candidates.remove(choice)

    return sorted(selected)


def build_samples(records: Sequence[Record], min_prefix_steps: int, max_prefix_samples_per_coin: int = 0) -> Tuple[List[Sample], FeatureStats]:
    samples: List[Sample] = []
    for record in records:
        start_ts = record.steps[0].timestamp_ms
        end_ts = record.steps[-1].timestamp_ms
        total_duration = record.metrics.total_duration_s
        shrinkable_radius = max(1.0, record.metrics.start_radius_m - record.metrics.floor_radius_m)
        for end in select_prefix_end_indices(record, min_prefix_steps, max_prefix_samples_per_coin):
            prefix = record.steps[:end]
            last = prefix[-1]
            east_m, north_m = signed_offset_m((last.lat, last.lng), (record.exact_lat, record.exact_lng))
            public_east_m, public_north_m = signed_offset_m(
                (last.lat, last.lng),
                (record.steps[-1].lat, record.steps[-1].lng),
            )
            residual_east_m, residual_north_m = signed_offset_m(
                (record.steps[-1].lat, record.steps[-1].lng),
                (record.exact_lat, record.exact_lng),
            )
            next_step_offset_m = (0.0, 0.0)
            next_radius_delta_norm = 0.0
            next_turn_flag = 0.0
            if end < len(record.steps):
                next_step = record.steps[end]
                next_step_offset_m = signed_offset_m((last.lat, last.lng), (next_step.lat, next_step.lng))
                next_radius_delta_norm = (last.radius_m - next_step.radius_m) / max(1.0, last.radius_m)
                if end >= 2:
                    prev_step = record.steps[end - 2]
                    prev_east, prev_north = signed_offset_m((prev_step.lat, prev_step.lng), (last.lat, last.lng))
                    curr_east, curr_north = next_step_offset_m
                    if math.hypot(prev_east, prev_north) > 1e-6 and math.hypot(curr_east, curr_north) > 1e-6:
                        prev_heading = math.atan2(prev_north, prev_east)
                        curr_heading = math.atan2(curr_north, curr_east)
                        next_turn_flag = 1.0 if abs(wrap_angle_rad(curr_heading - prev_heading)) > math.radians(26.0) else 0.0
            progress_to_floor = clamp(
                (record.metrics.start_radius_m - last.radius_m) / shrinkable_radius,
                0.0,
                1.0,
            )
            remaining_progress = max(0.0, 1.0 - progress_to_floor)
            elapsed_s = 0.0
            if start_ts is not None and last.timestamp_ms is not None and last.timestamp_ms >= start_ts:
                elapsed_s = (last.timestamp_ms - start_ts) / 1000.0
            remaining_duration_s = 0.0
            if end_ts is not None and last.timestamp_ms is not None and end_ts >= last.timestamp_ms:
                remaining_duration_s = (end_ts - last.timestamp_ms) / 1000.0
            profile = build_sequence_profile(prefix)
            samples.append(
                Sample(
                    sample_id=f"{record.entry_id}:{end}",
                    entry_id=record.entry_id,
                    room_code=record.room_code,
                    coin_label=record.coin_label,
                    prefix_steps=prefix,
                    prefix_len=end,
                    total_steps=len(record.steps),
                    output_offset_m=(east_m, north_m),
                    output_norm=(east_m / last.radius_m, north_m / last.radius_m),
                    progress_to_floor=progress_to_floor,
                    remaining_progress=remaining_progress,
                    total_duration_ratio=(elapsed_s / total_duration) if total_duration > 0 else progress_to_floor,
                    remaining_duration_s=remaining_duration_s,
                    profile=profile,
                    embedding=build_embedding_from_steps(prefix),
                    signature=build_signature_vector(prefix),
                    exact_coords=(record.exact_lat, record.exact_lng),
                    record_metrics=record.metrics,
                    family_id=record.family_id,
                    public_offset_m=(public_east_m, public_north_m),
                    residual_offset_m=(residual_east_m, residual_north_m),
                    next_step_offset_m=next_step_offset_m,
                    next_radius_delta_norm=next_radius_delta_norm,
                    next_turn_flag=next_turn_flag,
                    remaining_steps=max(0, len(record.steps) - end),
                )
            )
    feature_stats = build_feature_stats([sample.embedding for sample in samples])
    for sample in samples:
        sample.embedding = standardize_vector(sample.embedding, feature_stats)
    return samples, feature_stats


def get_indices() -> Dict[str, Optional[List[int]]]:
    stride = EMBED_LEN
    return {
        "full": None,
        "shape": [index for index in range(stride * 3)],
        "motion": [stride * 3 + index for index in range(stride * 6)],
    }


def profile_distance(query_profile: Dict[str, float], sample_profile: Dict[str, float]) -> float:
    keys = [
        "radiusRatio",
        "startRadiusKm",
        "currentRadiusKm",
        "shrinkFraction",
        "driftRatio",
        "meanStepDistanceRatio",
        "durationNorm",
        "shrinkRateNorm",
        "bundleCountRatio",
        "meanBundleRunRatio",
        "maxBundleRunRatio",
        "finalBundleRunRatio",
        "turnDensity",
        "headingChangeNorm",
        "finalHeadingEastNorm",
        "finalHeadingNorthNorm",
    ]
    total = 0.0
    for key in keys:
        total += (float(query_profile.get(key, 0.0)) - float(sample_profile.get(key, 0.0))) ** 2
    return math.sqrt(total)


def estimate_floor_state(observed_steps: Sequence[Step], neighbors: Sequence[Dict[str, Any]]) -> Dict[str, float]:
    start_radius = max(1.0, observed_steps[0].radius_m)
    current_radius = max(1.0, observed_steps[-1].radius_m)
    weights = [float(item["weight"]) for item in neighbors if math.isfinite(float(item["weight"]))]
    floor_ratios = [
        float(item["sample"].record_metrics.floor_radius_m) / max(1.0, float(item["sample"].record_metrics.start_radius_m))
        for item in neighbors
    ]
    sample_remaining = [max(0.02, float(item["sample"].remaining_progress)) for item in neighbors]
    sample_durations = [max(1.0, float(item["sample"].remaining_duration_s)) for item in neighbors]
    floor_ratio_mean = weighted_mean(floor_ratios, weights) if weights else mean(floor_ratios)
    est_floor_radius = clamp(start_radius * floor_ratio_mean, 1.0, current_radius)
    shrinkable = max(1.0, start_radius - est_floor_radius)
    remaining_progress = clamp((current_radius - est_floor_radius) / shrinkable, 0.0, 1.0)
    return {
        "estimatedFloorRadiusM": est_floor_radius,
        "remainingProgress": remaining_progress,
        "sampleRemainingProgressMean": weighted_mean(sample_remaining, weights) if weights else mean(sample_remaining),
        "sampleRemainingDurationMean": weighted_mean(sample_durations, weights) if weights else mean(sample_durations),
    }


def get_base_config_library() -> List[Dict[str, Any]]:
    return [
        {
            "id": "algo-balanced",
            "label": "Balanced",
            "penalties": {"radius": 3.2, "drift": 2.2, "cadence": 1.25, "size": 1.45, "duration": 1.75, "shrinkRate": 1.55, "signature": 0.95, "bundle": 1.15},
            "roomFactor": 0.84,
            "familyBoost": 0.26,
            "ensemble": {"fullRaw": 0.22, "fullScaled": 0.28, "shapeRaw": 0.15, "shapeScaled": 0.19, "motionRaw": 0.07, "motionScaled": 0.09},
            "blend": {"full": 0.52, "shape": 0.28, "motion": 0.20},
            "reverseBlend": {"direct": 0.34, "decomposed": 0.42, "generator": 0.24},
            "progressScaleBounds": [0.65, 1.45],
            "remainingScaleBounds": [0.55, 1.75],
        },
        {
            "id": "algo-shape",
            "label": "Shape-first",
            "penalties": {"radius": 4.8, "drift": 3.1, "cadence": 1.4, "size": 1.6, "duration": 2.05, "shrinkRate": 1.8, "signature": 1.10, "bundle": 1.35},
            "roomFactor": 0.86,
            "familyBoost": 0.34,
            "ensemble": {"fullRaw": 0.16, "fullScaled": 0.23, "shapeRaw": 0.20, "shapeScaled": 0.24, "motionRaw": 0.05, "motionScaled": 0.12},
            "blend": {"full": 0.35, "shape": 0.45, "motion": 0.20},
            "reverseBlend": {"direct": 0.28, "decomposed": 0.47, "generator": 0.25},
            "progressScaleBounds": [0.65, 1.45],
            "remainingScaleBounds": [0.55, 1.75],
        },
        {
            "id": "algo-room",
            "label": "Room-aware",
            "penalties": {"radius": 3.0, "drift": 2.0, "cadence": 1.1, "size": 1.35, "duration": 1.65, "shrinkRate": 1.5, "signature": 0.85, "bundle": 1.05},
            "roomFactor": 0.74,
            "familyBoost": 0.24,
            "ensemble": {"fullRaw": 0.20, "fullScaled": 0.27, "shapeRaw": 0.13, "shapeScaled": 0.18, "motionRaw": 0.08, "motionScaled": 0.14},
            "blend": {"full": 0.58, "shape": 0.22, "motion": 0.20},
            "reverseBlend": {"direct": 0.36, "decomposed": 0.41, "generator": 0.23},
            "progressScaleBounds": [0.65, 1.45],
            "remainingScaleBounds": [0.55, 1.75],
        },
        {
            "id": "algo-drift",
            "label": "Drift-first",
            "penalties": {"radius": 2.6, "drift": 3.8, "cadence": 1.55, "size": 1.2, "duration": 1.6, "shrinkRate": 1.9, "signature": 1.0, "bundle": 1.2},
            "roomFactor": 0.84,
            "familyBoost": 0.28,
            "ensemble": {"fullRaw": 0.24, "fullScaled": 0.23, "shapeRaw": 0.12, "shapeScaled": 0.16, "motionRaw": 0.10, "motionScaled": 0.15},
            "blend": {"full": 0.42, "shape": 0.22, "motion": 0.36},
            "reverseBlend": {"direct": 0.31, "decomposed": 0.36, "generator": 0.33},
            "progressScaleBounds": [0.65, 1.45],
            "remainingScaleBounds": [0.55, 1.75],
        },
    ]


def create_config(base: Dict[str, Any], suffix: str, overrides: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    overrides = overrides or {}
    ensemble = normalize_weight_map(overrides.get("ensemble", base["ensemble"]), base["ensemble"])
    blend = normalize_weight_map(overrides.get("blend", base["blend"]), base["blend"])
    reverse_blend = normalize_weight_map(overrides.get("reverseBlend", base["reverseBlend"]), base["reverseBlend"])
    config = {
        **base,
        **overrides,
        "id": f"{base['id']}-{suffix}",
        "label": f"{base['label']} {suffix}",
        "penalties": {**base["penalties"], **(overrides.get("penalties") or {})},
        "ensemble": ensemble,
        "blend": blend,
        "reverseBlend": reverse_blend,
        "progressScaleBounds": list(overrides.get("progressScaleBounds", base["progressScaleBounds"])),
        "remainingScaleBounds": list(overrides.get("remainingScaleBounds", base["remainingScaleBounds"])),
    }
    return config


def build_initial_candidates() -> List[Dict[str, Any]]:
    room_biases = [0.76, 0.90]
    duration_biases = [1.35, 1.60, 1.85]
    shrink_rate_biases = [1.4, 1.8]
    scale_modes = [
        {"suffix": "memory", "ensembleDelta": {"fullScaled": 0.08, "shapeScaled": 0.05, "fullRaw": -0.06, "motionRaw": -0.03}, "blendDelta": {"full": 0.04, "shape": 0.02, "motion": -0.06}},
        {"suffix": "balanced", "ensembleDelta": {}, "blendDelta": {}},
    ]
    candidates: List[Dict[str, Any]] = []
    for base in get_base_config_library():
        for room_factor in room_biases:
            for duration_weight in duration_biases:
                for shrink_rate_weight in shrink_rate_biases:
                    for scale_mode in scale_modes:
                        ensemble = shift_weights(base["ensemble"], scale_mode["ensembleDelta"], base["ensemble"])
                        blend = shift_weights(base["blend"], scale_mode["blendDelta"], base["blend"])
                        candidates.append(create_config(
                            base,
                            f"{scale_mode['suffix']}-rf{round(room_factor * 100)}-du{round(duration_weight * 100)}-sr{round(shrink_rate_weight * 100)}",
                            {
                                "roomFactor": room_factor,
                                "penalties": {"duration": duration_weight, "shrinkRate": shrink_rate_weight},
                                "ensemble": ensemble,
                                "blend": blend,
                            },
                        ))
    return candidates


def config_fingerprint(config: Dict[str, Any]) -> str:
    return json.dumps({
        "roomFactor": round(float(config["roomFactor"]), 5),
        "familyBoost": round(float(config.get("familyBoost", 0.28)), 5),
        "penalties": {key: round(float(value), 5) for key, value in sorted(config["penalties"].items())},
        "ensemble": {key: round(float(value), 6) for key, value in sorted(config["ensemble"].items())},
        "blend": {key: round(float(value), 6) for key, value in sorted(config["blend"].items())},
        "reverseBlend": {key: round(float(value), 6) for key, value in sorted(config["reverseBlend"].items())},
        "progressScaleBounds": [round(float(value), 5) for value in config["progressScaleBounds"]],
        "remainingScaleBounds": [round(float(value), 5) for value in config["remainingScaleBounds"]],
    }, sort_keys=True)


def mutate_config(base: Dict[str, Any], rng: random.Random) -> Dict[str, Any]:
    penalties = {}
    for key, value in base["penalties"].items():
        penalties[key] = max(0.3, float(value) * (1 + rng.uniform(-0.16, 0.16)))
    ensemble = shift_weights(base["ensemble"], {key: rng.uniform(-0.05, 0.05) for key in base["ensemble"]}, base["ensemble"])
    blend = shift_weights(base["blend"], {key: rng.uniform(-0.06, 0.06) for key in base["blend"]}, base["blend"])
    reverse_blend = shift_weights(base["reverseBlend"], {key: rng.uniform(-0.06, 0.06) for key in base["reverseBlend"]}, base["reverseBlend"])
    progress_bounds = list(base["progressScaleBounds"])
    remaining_bounds = list(base["remainingScaleBounds"])
    progress_bounds[0] = clamp(progress_bounds[0] + rng.uniform(-0.07, 0.05), 0.35, 1.1)
    progress_bounds[1] = clamp(progress_bounds[1] + rng.uniform(-0.08, 0.08), max(progress_bounds[0] + 0.05, 1.05), 2.2)
    remaining_bounds[0] = clamp(remaining_bounds[0] + rng.uniform(-0.08, 0.08), 0.25, 1.05)
    remaining_bounds[1] = clamp(remaining_bounds[1] + rng.uniform(-0.12, 0.12), max(remaining_bounds[0] + 0.12, 1.0), 2.6)
    return create_config(
        base,
        f"mut{rng.randint(10_000, 99_999)}",
        {
            "roomFactor": clamp(float(base["roomFactor"]) + rng.uniform(-0.06, 0.06), 0.58, 0.98),
            "familyBoost": clamp(float(base.get("familyBoost", 0.28)) + rng.uniform(-0.08, 0.08), 0.0, 0.55),
            "penalties": penalties,
            "ensemble": ensemble,
            "blend": blend,
            "reverseBlend": reverse_blend,
            "progressScaleBounds": progress_bounds,
            "remainingScaleBounds": remaining_bounds,
        },
    )


def predict_neighbors(model: Dict[str, Any], query_steps: Sequence[Step], query_profile: Dict[str, float], query_signature: Sequence[float], query_embedding: Sequence[float], config: Dict[str, Any], indices: Optional[Sequence[int]], exclude_entry_id: str = "", room_code: str = "", limit: int = K_NEIGHBORS) -> Dict[str, Any]:
    scored = []
    penalties = config["penalties"]
    room_factor = clamp(float(config["roomFactor"]), 0.55, 1.0)
    family_guess = classify_family(query_profile, model["familyModel"])
    family_boost = clamp(float(config.get("familyBoost", 0.28)), 0.0, 0.6)
    for sample in model["samples"]:
        if exclude_entry_id and sample.entry_id == exclude_entry_id:
            continue
        embed_distance = distance(query_embedding, sample.embedding, indices)
        radius_penalty = abs(float(query_profile.get("radiusRatio", 0.0)) - float(sample.profile.get("radiusRatio", 0.0))) * float(penalties["radius"])
        drift_penalty = abs(float(query_profile.get("driftRatio", 0.0)) - float(sample.profile.get("driftRatio", 0.0))) * float(penalties["drift"])
        cadence_penalty = abs(float(query_profile.get("meanStepDistanceRatio", 0.0)) - float(sample.profile.get("meanStepDistanceRatio", 0.0))) * float(penalties["cadence"])
        size_penalty = abs(float(query_profile.get("stepCount", 0.0)) - float(sample.profile.get("stepCount", 0.0))) / 10.0 * float(penalties["size"])
        duration_penalty = abs(float(query_profile.get("durationNorm", 0.0)) - float(sample.profile.get("durationNorm", 0.0))) * float(penalties["duration"])
        shrink_penalty = abs(float(query_profile.get("shrinkRateNorm", 0.0)) - float(sample.profile.get("shrinkRateNorm", 0.0))) * float(penalties["shrinkRate"])
        bundle_penalty = (
            abs(float(query_profile.get("bundleCountRatio", 0.0)) - float(sample.profile.get("bundleCountRatio", 0.0))) * 0.7
            + abs(float(query_profile.get("meanBundleRunRatio", 0.0)) - float(sample.profile.get("meanBundleRunRatio", 0.0))) * 1.0
            + abs(float(query_profile.get("maxBundleRunRatio", 0.0)) - float(sample.profile.get("maxBundleRunRatio", 0.0))) * 0.8
            + abs(float(query_profile.get("finalBundleRunRatio", 0.0)) - float(sample.profile.get("finalBundleRunRatio", 0.0))) * 1.2
            + abs(float(query_profile.get("turnDensity", 0.0)) - float(sample.profile.get("turnDensity", 0.0))) * 0.85
            + abs(float(query_profile.get("headingChangeNorm", 0.0)) - float(sample.profile.get("headingChangeNorm", 0.0))) * 0.7
            + abs(float(query_profile.get("finalHeadingEastNorm", 0.0)) - float(sample.profile.get("finalHeadingEastNorm", 0.0))) * 0.45
            + abs(float(query_profile.get("finalHeadingNorthNorm", 0.0)) - float(sample.profile.get("finalHeadingNorthNorm", 0.0))) * 0.45
        ) * float(penalties.get("bundle", 1.0))
        signature_penalty = distance(query_signature, sample.signature) * float(penalties.get("signature", 0.95))
        total = embed_distance + radius_penalty + drift_penalty + cadence_penalty + size_penalty + duration_penalty + shrink_penalty + bundle_penalty + signature_penalty
        if room_code and sample.room_code and sample.room_code == room_code:
            total *= room_factor
        if sample.family_id:
            same_family_score = float(family_guess["scores"].get(sample.family_id, 0.0))
            total *= clamp(1.0 - (same_family_score * family_boost), 0.58, 1.08)
        scored.append((total, sample))
    scored.sort(key=lambda item: item[0])
    neighbors = scored[: max(1, limit)]
    if not neighbors:
        return {
            "neighbors": [],
            "prediction": (0.0, 0.0),
            "predictionScaled": (0.0, 0.0),
            "publicPrediction": (0.0, 0.0),
            "publicPredictionScaled": (0.0, 0.0),
            "residualPrediction": (0.0, 0.0),
            "nextStepPrediction": (0.0, 0.0),
            "nextRadiusDeltaNorm": 0.0,
            "remainingStepsEstimate": 0.0,
            "bundlePersistence": 0.0,
            "spread": 0.0,
            "familyGuess": family_guess,
        }
    weights = [1 / max(1e-6, item[0] + 0.08) for item in neighbors]
    raw_offsets = [item[1].output_offset_m for item in neighbors]
    prediction = (
        weighted_mean([offset[0] for offset in raw_offsets], weights),
        weighted_mean([offset[1] for offset in raw_offsets], weights),
    )
    cloud = [{"distance": dist, "sample": sample, "weight": weight} for (dist, sample), weight in zip(neighbors, weights)]
    floor_state = estimate_floor_state(query_steps, cloud)
    progress_ratio = (floor_state["remainingProgress"] / max(0.02, floor_state["sampleRemainingProgressMean"]))
    progress_scale = clamp(progress_ratio, float(config["progressScaleBounds"][0]), float(config["progressScaleBounds"][1]))
    scaled_offsets = []
    public_offsets = []
    public_scaled_offsets = []
    for item in cloud:
        sample_remaining = max(0.02, float(item["sample"].remaining_progress))
        per_neighbor_scale = clamp(
            floor_state["remainingProgress"] / sample_remaining,
            float(config["remainingScaleBounds"][0]),
            float(config["remainingScaleBounds"][1]),
        )
        east, north = item["sample"].output_offset_m
        scaled_offsets.append((east * per_neighbor_scale, north * per_neighbor_scale))
        item["scaled_offset"] = scaled_offsets[-1]
        public_east, public_north = item["sample"].public_offset_m
        public_offsets.append((public_east, public_north))
        public_scaled_offsets.append((public_east * per_neighbor_scale, public_north * per_neighbor_scale))
        item["public_scaled_offset"] = public_scaled_offsets[-1]
    prediction_scaled = (
        weighted_mean([offset[0] for offset in scaled_offsets], weights),
        weighted_mean([offset[1] for offset in scaled_offsets], weights),
    )
    public_prediction = (
        weighted_mean([offset[0] for offset in public_offsets], weights),
        weighted_mean([offset[1] for offset in public_offsets], weights),
    )
    public_prediction_scaled = (
        weighted_mean([offset[0] for offset in public_scaled_offsets], weights),
        weighted_mean([offset[1] for offset in public_scaled_offsets], weights),
    )
    residual_prediction = (
        weighted_mean([item["sample"].residual_offset_m[0] for item in cloud], weights),
        weighted_mean([item["sample"].residual_offset_m[1] for item in cloud], weights),
    )
    next_step_prediction = (
        weighted_mean([item["sample"].next_step_offset_m[0] for item in cloud], weights),
        weighted_mean([item["sample"].next_step_offset_m[1] for item in cloud], weights),
    )
    next_radius_delta_norm = weighted_mean([item["sample"].next_radius_delta_norm for item in cloud], weights)
    remaining_steps_est = weighted_mean([float(item["sample"].remaining_steps) for item in cloud], weights)
    bundle_persistence = weighted_mean([1.0 - float(item["sample"].next_turn_flag) for item in cloud], weights)
    spread = math.sqrt(mean([
        ((offset[0] - prediction[0]) ** 2) + ((offset[1] - prediction[1]) ** 2)
        for offset in raw_offsets
    ]))
    return {
        "neighbors": cloud,
        "prediction": prediction,
        "predictionScaled": prediction_scaled,
        "publicPrediction": public_prediction,
        "publicPredictionScaled": public_prediction_scaled,
        "residualPrediction": residual_prediction,
        "nextStepPrediction": next_step_prediction,
        "nextRadiusDeltaNorm": next_radius_delta_norm,
        "remainingStepsEstimate": remaining_steps_est,
        "bundlePersistence": bundle_persistence,
        "spread": spread,
        "floorState": floor_state,
        "progressScale": progress_scale,
        "familyGuess": family_guess,
    }


def solve_core(model: Dict[str, Any], query_sample: Sample, config: Dict[str, Any], exclude_entry_id: str = "") -> Dict[str, Any]:
    query_profile = query_sample.profile
    query_signature = query_sample.signature
    query_embedding = query_sample.embedding
    room_code = query_sample.room_code
    full_model = predict_neighbors(model, query_sample.prefix_steps, query_profile, query_signature, query_embedding, config, model["indices"]["full"], exclude_entry_id, room_code, K_NEIGHBORS)
    shape_model = predict_neighbors(model, query_sample.prefix_steps, query_profile, query_signature, query_embedding, config, model["indices"]["shape"], exclude_entry_id, room_code, max(6, math.ceil(K_NEIGHBORS * 0.8)))
    motion_model = predict_neighbors(model, query_sample.prefix_steps, query_profile, query_signature, query_embedding, config, model["indices"]["motion"], exclude_entry_id, room_code, max(6, math.ceil(K_NEIGHBORS * 0.8)))
    ensemble = config["ensemble"]
    predictions = [
        (float(ensemble["fullRaw"]), full_model["prediction"]),
        (float(ensemble["fullScaled"]), full_model["predictionScaled"]),
        (float(ensemble["shapeRaw"]), shape_model["prediction"]),
        (float(ensemble["shapeScaled"]), shape_model["predictionScaled"]),
        (float(ensemble["motionRaw"]), motion_model["prediction"]),
        (float(ensemble["motionScaled"]), motion_model["predictionScaled"]),
    ]
    positive = [(weight, offset) for weight, offset in predictions if weight > 0]
    total_weight = sum(weight for weight, _ in positive) or 1.0
    mean_offset = (
        sum(offset[0] * weight for weight, offset in positive) / total_weight,
        sum(offset[1] * weight for weight, offset in positive) / total_weight,
    )
    model_std = math.sqrt(mean([
        ((offset[0] - mean_offset[0]) ** 2) + ((offset[1] - mean_offset[1]) ** 2)
        for _, offset in positive
    ]))

    blend = config["blend"]
    reverse_blend = config.get("reverseBlend") or {"direct": 0.34, "decomposed": 0.42, "generator": 0.24}

    def combine_component(key: str) -> Tuple[float, float]:
        weighted_offsets = []
        for weight, neighbor_model in [
            (float(blend["full"]), full_model),
            (float(blend["shape"]), shape_model),
            (float(blend["motion"]), motion_model),
        ]:
            component = neighbor_model.get(key) or (0.0, 0.0)
            weighted_offsets.append((weight, component))
        total = sum(weight for weight, _ in weighted_offsets) or 1.0
        return (
            sum(weight * float(offset[0]) for weight, offset in weighted_offsets) / total,
            sum(weight * float(offset[1]) for weight, offset in weighted_offsets) / total,
        )

    public_offset = combine_component("publicPredictionScaled")
    residual_offset = combine_component("residualPrediction")
    next_step_offset = combine_component("nextStepPrediction")
    remaining_steps_est = weighted_mean(
        [
            float(full_model.get("remainingStepsEstimate", 0.0)),
            float(shape_model.get("remainingStepsEstimate", 0.0)),
            float(motion_model.get("remainingStepsEstimate", 0.0)),
        ],
        [float(blend["full"]), float(blend["shape"]), float(blend["motion"])],
    )
    bundle_persistence = weighted_mean(
        [
            float(full_model.get("bundlePersistence", 0.0)),
            float(shape_model.get("bundlePersistence", 0.0)),
            float(motion_model.get("bundlePersistence", 0.0)),
        ],
        [float(blend["full"]), float(blend["shape"]), float(blend["motion"])],
    )
    generator_public_offset = (
        next_step_offset[0] * remaining_steps_est * clamp(0.55 + (bundle_persistence * 0.45), 0.45, 1.25),
        next_step_offset[1] * remaining_steps_est * clamp(0.55 + (bundle_persistence * 0.45), 0.45, 1.25),
    )
    decomposed_exact_offset = (
        public_offset[0] + residual_offset[0],
        public_offset[1] + residual_offset[1],
    )
    generator_exact_offset = (
        generator_public_offset[0] + residual_offset[0],
        generator_public_offset[1] + residual_offset[1],
    )

    blended_neighbors = []
    for weight, neighbor_model in [
        (float(blend["full"]), full_model),
        (float(blend["shape"]), shape_model),
        (float(blend["motion"]), motion_model),
    ]:
        for item in neighbor_model["neighbors"]:
            blended_neighbors.append({
                "sample": item["sample"],
                "weight": weight * float(item["weight"]),
                "scaled_offset": item.get("scaled_offset", item["sample"].output_offset_m),
            })
    if not blended_neighbors:
        return {"predictedOffsetM": (0.0, 0.0), "uncertaintyM": float("inf")}
    total_neighbor_weight = sum(item["weight"] for item in blended_neighbors) or 1.0
    direct_neighbor_offset = (
        sum(item["scaled_offset"][0] * item["weight"] for item in blended_neighbors) / total_neighbor_weight,
        sum(item["scaled_offset"][1] * item["weight"] for item in blended_neighbors) / total_neighbor_weight,
    )
    predicted_offset = (
        float(reverse_blend["direct"]) * direct_neighbor_offset[0]
        + float(reverse_blend["decomposed"]) * decomposed_exact_offset[0]
        + float(reverse_blend["generator"]) * generator_exact_offset[0],
        float(reverse_blend["direct"]) * direct_neighbor_offset[1]
        + float(reverse_blend["decomposed"]) * decomposed_exact_offset[1]
        + float(reverse_blend["generator"]) * generator_exact_offset[1],
    )
    direct_neighbor_spread = math.sqrt(mean([
        ((item["scaled_offset"][0] - predicted_offset[0]) ** 2) + ((item["scaled_offset"][1] - predicted_offset[1]) ** 2)
        for item in blended_neighbors
    ]))
    hypothesis_offsets = [mean_offset, direct_neighbor_offset, decomposed_exact_offset, generator_exact_offset, predicted_offset]
    hypothesis_spread = math.sqrt(mean([
        ((offset[0] - predicted_offset[0]) ** 2) + ((offset[1] - predicted_offset[1]) ** 2)
        for offset in hypothesis_offsets
    ]))
    family_confidence = max(
        float(full_model.get("familyGuess", {}).get("scores", {}).get(full_model.get("familyGuess", {}).get("familyId", ""), 0.0)),
        float(shape_model.get("familyGuess", {}).get("scores", {}).get(shape_model.get("familyGuess", {}).get("familyId", ""), 0.0)),
        float(motion_model.get("familyGuess", {}).get("scores", {}).get(motion_model.get("familyGuess", {}).get("familyId", ""), 0.0)),
        0.05,
    )
    uncertainty_m = max(
        12.0,
        direct_neighbor_spread
        + full_model["spread"]
        + model_std * max(0.65, full_model.get("progressScale", 1.0))
        + hypothesis_spread * (1.15 - family_confidence),
    )
    return {
        "predictedOffsetM": predicted_offset,
        "uncertaintyM": uncertainty_m,
        "familyId": max(
            [
                full_model.get("familyGuess", {}).get("familyId", ""),
                shape_model.get("familyGuess", {}).get("familyId", ""),
                motion_model.get("familyGuess", {}).get("familyId", ""),
            ],
            key=lambda family_id: (
                float(full_model.get("familyGuess", {}).get("scores", {}).get(family_id, 0.0))
                + float(shape_model.get("familyGuess", {}).get("scores", {}).get(family_id, 0.0))
                + float(motion_model.get("familyGuess", {}).get("scores", {}).get(family_id, 0.0))
            ),
        ),
        "components": {
            "directNeighbor": direct_neighbor_offset,
            "decomposed": decomposed_exact_offset,
            "generator": generator_exact_offset,
            "publicOffset": public_offset,
            "residualOffset": residual_offset,
            "generatorPublicOffset": generator_public_offset,
            "remainingStepsEstimate": remaining_steps_est,
            "bundlePersistence": bundle_persistence,
        },
    }


def score_metrics(metrics: Dict[str, Any], target_error_m: float, target_conf_m: float) -> float:
    fail_count = int(metrics["failCount"])
    fail_error_excess = float(metrics["failErrorExcessM"])
    fail_conf_excess = float(metrics["failConfidenceExcessM"])
    return (
        fail_count * 1_000_000.0
        + fail_error_excess * 2_000.0
        + fail_conf_excess * 1_500.0
        + float(metrics["meanErrorM"]) * 20.0
        + float(metrics["meanConfidenceM"]) * 6.0
        + float(metrics["p70ErrorM"]) * 10.0
        + float(metrics["maxErrorM"]) * 8.0
        + float(metrics["maxConfidenceM"]) * 5.0
        - (1000.0 if metrics["allWithinTarget"] else 0.0)
    )


def backtest_config(model: Dict[str, Any], config: Dict[str, Any], target_error_m: float, target_conf_m: float, abort_score: Optional[float] = None) -> Dict[str, Any]:
    sample_errors: List[float] = []
    sample_confidences: List[float] = []
    per_coin: Dict[str, List[Tuple[float, float]]] = {}
    fail_count = 0
    fail_error_excess = 0.0
    fail_conf_excess = 0.0
    for sample in model["samples"]:
        solved = solve_core(model, sample, config, exclude_entry_id=sample.entry_id)
        predicted_offset = solved["predictedOffsetM"]
        error_m = math.hypot(
            predicted_offset[0] - sample.output_offset_m[0],
            predicted_offset[1] - sample.output_offset_m[1],
        )
        confidence_m = float(solved["uncertaintyM"])
        sample_errors.append(error_m)
        sample_confidences.append(confidence_m)
        per_coin.setdefault(sample.entry_id, []).append((error_m, confidence_m))
        if error_m > target_error_m or confidence_m > target_conf_m:
            fail_count += 1
            fail_error_excess += max(0.0, error_m - target_error_m)
            fail_conf_excess += max(0.0, confidence_m - target_conf_m)
        if abort_score is not None and len(sample_errors) >= 24:
            running = {
                "failCount": fail_count,
                "failErrorExcessM": fail_error_excess,
                "failConfidenceExcessM": fail_conf_excess,
                "meanErrorM": mean(sample_errors),
                "meanConfidenceM": mean(sample_confidences),
                "p70ErrorM": weighted_percentile(sample_errors, [1.0] * len(sample_errors), 0.7),
                "maxErrorM": max(sample_errors),
                "maxConfidenceM": max(sample_confidences),
                "allWithinTarget": False,
            }
            running["score"] = score_metrics(running, target_error_m, target_conf_m)
            if running["score"] > abort_score * 1.025:
                running["abortedEarly"] = True
                return running
    per_coin_error = [mean([item[0] for item in values]) for values in per_coin.values()]
    per_coin_conf = [mean([item[1] for item in values]) for values in per_coin.values()]
    metrics = {
        "sampleCount": len(sample_errors),
        "coinCount": len(per_coin),
        "meanErrorM": mean(sample_errors),
        "medianErrorM": weighted_percentile(sample_errors, [1.0] * len(sample_errors), 0.5),
        "p70ErrorM": weighted_percentile(sample_errors, [1.0] * len(sample_errors), 0.7),
        "maxErrorM": max(sample_errors) if sample_errors else float("inf"),
        "meanConfidenceM": mean(sample_confidences),
        "medianConfidenceM": weighted_percentile(sample_confidences, [1.0] * len(sample_confidences), 0.5),
        "maxConfidenceM": max(sample_confidences) if sample_confidences else float("inf"),
        "coinMeanErrorM": mean(per_coin_error),
        "coinMeanConfidenceM": mean(per_coin_conf),
        "failCount": fail_count,
        "failErrorExcessM": fail_error_excess,
        "failConfidenceExcessM": fail_conf_excess,
        "allWithinTarget": fail_count == 0,
        "abortedEarly": False,
    }
    metrics["score"] = score_metrics(metrics, target_error_m, target_conf_m)
    return metrics


def formula_string(config: Dict[str, Any]) -> str:
    penalties = config["penalties"]
    reverse_blend = config.get("reverseBlend") or {}
    return (
        f"roomFactor={config['roomFactor']:.3f}; "
        f"familyBoost={float(config.get('familyBoost', 0.28)):.3f}; "
        f"penalties(radius={penalties['radius']:.3f}, drift={penalties['drift']:.3f}, cadence={penalties['cadence']:.3f}, "
        f"size={penalties['size']:.3f}, duration={penalties['duration']:.3f}, shrinkRate={penalties['shrinkRate']:.3f}, signature={penalties['signature']:.3f}, bundle={penalties['bundle']:.3f}); "
        f"reverseBlend(direct={float(reverse_blend.get('direct', 0.34)):.3f}, decomposed={float(reverse_blend.get('decomposed', 0.42)):.3f}, generator={float(reverse_blend.get('generator', 0.24)):.3f}); "
        f"progressScale=[{config['progressScaleBounds'][0]:.3f},{config['progressScaleBounds'][1]:.3f}]; "
        f"remainingScale=[{config['remainingScaleBounds'][0]:.3f},{config['remainingScaleBounds'][1]:.3f}]"
    )


def search_best_config(model: Dict[str, Any], args: argparse.Namespace) -> Tuple[Dict[str, Any], Dict[str, Any], List[Dict[str, Any]]]:
    rng = random.Random(args.seed)
    seen: set[str] = set()
    leaderboard: List[Dict[str, Any]] = []
    start_time = time.perf_counter()
    stale_iterations = 0
    iterations = 0
    best_config: Optional[Dict[str, Any]] = None
    best_metrics: Optional[Dict[str, Any]] = None

    def consider(config: Dict[str, Any]) -> None:
        nonlocal stale_iterations, iterations, best_config, best_metrics, leaderboard
        key = config_fingerprint(config)
        if key in seen:
            return
        seen.add(key)
        abort_score = float(best_metrics["score"]) if best_metrics else None
        metrics = backtest_config(model, config, args.target_error_m, args.target_confidence_m, abort_score)
        iterations += 1
        improved = best_metrics is None or float(metrics["score"]) < float(best_metrics["score"])
        if improved:
            best_config = config
            best_metrics = metrics
            stale_iterations = 0
        else:
            stale_iterations += 1
        leaderboard.append({"config": config, "metrics": metrics})
        leaderboard.sort(key=lambda item: float(item["metrics"]["score"]))
        del leaderboard[12:]
        if iterations % 10 == 0 or improved:
            status = "IMPROVED" if improved else "checked"
            print(
                f"[{iterations:04d}] {status}: fail={metrics['failCount']} "
                f"mean={metrics['meanErrorM']:.1f}m conf={metrics['meanConfidenceM']:.1f}m "
                f"max={metrics['maxErrorM']:.1f}m score={metrics['score']:.1f}"
            )

    for config in build_initial_candidates():
        if (time.perf_counter() - start_time) >= args.time_budget_sec or iterations >= args.max_iterations:
            break
        consider(config)
        if best_metrics and best_metrics["allWithinTarget"]:
            break

    while iterations < args.max_iterations and (time.perf_counter() - start_time) < args.time_budget_sec:
        if best_metrics and best_metrics["allWithinTarget"]:
            break
        if stale_iterations >= args.max_stale_iterations and leaderboard:
            break
        seed_pool = leaderboard[: min(5, len(leaderboard))] or [{"config": get_base_config_library()[0]}]
        parent = rng.choice(seed_pool)["config"]
        consider(mutate_config(parent, rng))

    if not best_config or not best_metrics:
        raise RuntimeError("No valid config was evaluated.")
    return best_config, best_metrics, leaderboard


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=False), encoding="utf-8")


def main() -> None:
    args = parse_args()
    source_meta: Dict[str, Any]
    excel_path = Path(args.excel_path).expanduser() if args.excel_path else None
    if excel_path:
        records = load_excel_records(excel_path, args.min_prefix_steps)
        source_meta = {
            "type": "excel",
            "path": str(excel_path),
            "sheet": "Shrinks",
        }
    else:
        load_env_file(Path(args.env_file))
        supabase_url = require_env("SUPABASE_URL")
        supabase_key = require_env("SUPABASE_ANON_KEY")
        raw_entries = fetch_exact_records(supabase_url, supabase_key, args.limit, args.room.strip().lower())
        records = build_records(raw_entries, args.min_prefix_steps)
        source_meta = {
            "type": "supabase",
            "table": API_TABLE,
            "room": args.room.strip().lower(),
        }
    family_model = build_family_model(records, args.seed)
    samples, feature_stats = build_samples(records, args.min_prefix_steps, args.max_prefix_samples_per_coin)
    if not records or not samples:
        raise RuntimeError("No exact-ended records with enough shrink steps were found.")

    print(f"Loaded {len(records)} exact-ended coins and {len(samples)} prefix samples.")
    model = {
        "records": records,
        "samples": samples,
        "feature_stats": feature_stats,
        "indices": get_indices(),
        "familyModel": family_model,
    }

    started_at = time.perf_counter()
    best_config, best_metrics, leaderboard = search_best_config(model, args)
    elapsed = time.perf_counter() - started_at

    formula_payload = {
        "version": FORMULA_VERSION,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "source": {
            **source_meta,
            "exactEndedCoinCount": len(records),
            "prefixSampleCount": len(samples),
        },
        "target": {
            "errorMeters": args.target_error_m,
            "confidenceMeters": args.target_confidence_m,
        },
        "search": {
            "timeBudgetSec": args.time_budget_sec,
            "maxIterations": args.max_iterations,
            "seed": args.seed,
            "elapsedSec": elapsed,
        },
        "config": best_config,
        "formulaString": formula_string(best_config),
        "fitMetrics": best_metrics,
        "familyModel": family_model,
    }
    report_payload = {
        **formula_payload,
        "leaderboard": [
            {
                "rank": index + 1,
                "configId": item["config"]["id"],
                "formulaString": formula_string(item["config"]),
                "metrics": item["metrics"],
            }
            for index, item in enumerate(leaderboard[:10])
        ],
        "note": (
            "The fitter searches until it either finds a config whose backtests keep every prefix sample "
            f"within {args.target_error_m:.0f}m error and {args.target_confidence_m:.0f}m confidence, "
            "or it hits the configured time/iteration budget and exports the best config found so far."
        ),
    }

    write_json(Path(args.output_formula), formula_payload)
    write_json(Path(args.output_report), report_payload)
    print(f"Wrote formula to {args.output_formula}")
    print(f"Wrote report to {args.output_report}")
    if best_metrics["allWithinTarget"]:
        print("Target achieved: all backtested samples stayed within the requested thresholds.")
    else:
        print(
            "Best-so-far exported. "
            f"{best_metrics['failCount']} samples still exceeded the requested thresholds."
        )


if __name__ == "__main__":
    main()
