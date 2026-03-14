# ===========================================================
# excel_logger.py — Log every shrink event to an Excel file
# ===========================================================
# Imported by main.py:  from excel_logger import log_shrink, set_true_location
#
# Uses openpyxl to maintain a structured spreadsheet with one sheet
# per coin (keyed by brand_name + coin_number or coin_id).
# ===========================================================

import os
import threading
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter

_lock = threading.Lock()
_last_logged = {}   # coin_id -> (lat, lng, radius)  — dedup guard


def _safe_sheet_name(raw: str) -> str:
    """Excel sheet names max 31 chars, no special chars."""
    name = raw.replace("/", "-").replace("\\", "-").replace(":", "-")
    name = name.replace("?", "").replace("*", "").replace("[", "(").replace("]", ")")
    return name[:31]


def _get_or_create_wb(xlsx_path: str):
    if os.path.exists(xlsx_path):
        return openpyxl.load_workbook(xlsx_path)
    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def _ensure_sheet(wb, sheet_name: str):
    safe = _safe_sheet_name(sheet_name)
    if safe not in wb.sheetnames:
        ws = wb.create_sheet(title=safe)
        headers = ["Timestamp", "Status", "Latitude", "Longitude", "Radius (m)",
                    "Brand", "Coin #", "Reward", "Is Smallest", "True Lat", "True Lng"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
            ws.column_dimensions[get_column_letter(col_idx)].width = 16
        return ws
    return wb[safe]


def log_shrink(item, xlsx_path="shrink_log.xlsx",
               status_filter=("ongoing", "verifying"),
               only_when_changed=True):
    """
    Log a single coin item's shrink data to an Excel file.

    Parameters
    ----------
    item : dict
        A coin payload dict from the Sqkii API, containing keys like
        coin_id, status, freeCircle, brand_name, coin_number, reward, etc.
    xlsx_path : str
        Path to the Excel workbook.
    status_filter : tuple
        Only log items whose status is in this tuple.
    only_when_changed : bool
        If True, skip logging if lat/lng/radius haven't changed since last call.
    """
    try:
        status = item.get("status", "")
        if status_filter and status not in status_filter:
            return

        coin_id = str(item.get("coin_id", "unknown"))
        brand = item.get("brand_name", "")
        coin_num = item.get("coin_number", "")
        reward = item.get("reward", "")
        is_smallest = item.get("is_smallest_public_circle", False)

        fc = item.get("freeCircle") or {}
        center = fc.get("center") or {}
        lat = center.get("lat")
        lng = center.get("lng")
        radius = fc.get("radius")

        # Dedup: skip if nothing changed
        if only_when_changed:
            key = (lat, lng, radius)
            if _last_logged.get(coin_id) == key:
                return
            _last_logged[coin_id] = key

        sheet_label = f"{brand} {coin_num}".strip() if brand else coin_id

        with _lock:
            wb = _get_or_create_wb(xlsx_path)
            ws = _ensure_sheet(wb, sheet_label)
            row = ws.max_row + 1
            ws.cell(row=row, column=1, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=row, column=2, value=status)
            ws.cell(row=row, column=3, value=lat)
            ws.cell(row=row, column=4, value=lng)
            ws.cell(row=row, column=5, value=radius)
            ws.cell(row=row, column=6, value=brand)
            ws.cell(row=row, column=7, value=coin_num)
            ws.cell(row=row, column=8, value=str(reward) if reward else "")
            ws.cell(row=row, column=9, value="Yes" if is_smallest else "")
            wb.save(xlsx_path)

    except Exception as e:
        print(f"[excel_logger] log_shrink error: {e}")


def set_true_location(coin_id, lat, lng, xlsx_path="shrink_log.xlsx"):
    """
    Update the 'True Lat' and 'True Lng' columns in the most recent row
    for a given coin, after the actual location has been confirmed.
    """
    try:
        if not os.path.exists(xlsx_path):
            return

        with _lock:
            wb = openpyxl.load_workbook(xlsx_path)
            # Search all sheets for matching coin_id
            for ws in wb.worksheets:
                last_row = ws.max_row
                if last_row < 2:
                    continue
                # Write true location to last row's True Lat/Lng columns (10, 11)
                ws.cell(row=last_row, column=10, value=lat)
                ws.cell(row=last_row, column=11, value=lng)
            wb.save(xlsx_path)

    except Exception as e:
        print(f"[excel_logger] set_true_location error: {e}")
